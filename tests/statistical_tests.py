import unittest
import numpy as np
import scipy.stats as stats
import statsmodels.stats.multicomp as mc
import sys
import os

# Add the project directory to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

class TestOneWayANOVA(unittest.TestCase):
    
    def setUp(self):
        # Sample normal datasets with known means and variances
        self.g1 = [1.2, 1.5, 1.8, 1.4, 1.6]
        self.g2 = [2.2, 2.5, 2.8, 2.4, 2.6]
        self.g3 = [3.2, 3.5, 3.8, 3.4, 3.6]
        self.all_groups = [self.g1, self.g2, self.g3]

    def test_standard_anova_calculation(self):
        # Perform standard ANOVA calculations manually/separately
        f_stat_expected, p_val_expected = stats.f_oneway(self.g1, self.g2, self.g3)
        
        # calculate sum of squares
        all_vals = self.g1 + self.g2 + self.g3
        grand_mean = np.mean(all_vals)
        ss_total = sum((x - grand_mean)**2 for x in all_vals)
        ss_between = len(self.g1)*(np.mean(self.g1) - grand_mean)**2 + \
                     len(self.g2)*(np.mean(self.g2) - grand_mean)**2 + \
                     len(self.g3)*(np.mean(self.g3) - grand_mean)**2
        ss_within = ss_total - ss_between
        
        df_between = 3 - 1
        df_within = 15 - 3
        
        ms_between = ss_between / df_between
        ms_within = ss_within / df_within
        f_stat = ms_between / ms_within
        
        self.assertAlmostEqual(f_stat, f_stat_expected)
        self.assertAlmostEqual(ss_total, ss_between + ss_within)
        self.assertEqual(df_between, 2)
        self.assertEqual(df_within, 12)

    def test_shapiro_wilk(self):
        # Shapiro-Wilk test on normal groups
        for idx, group in enumerate(self.all_groups):
            sh_stat, sh_p = stats.shapiro(group)
            self.assertTrue(sh_stat > 0)
            self.assertTrue(0 <= sh_p <= 1)

    def test_levene_test(self):
        # Levene's test on equal variances
        lev_stat, lev_p = stats.levene(self.g1, self.g2, self.g3)
        self.assertTrue(lev_stat >= 0)
        self.assertTrue(0 <= lev_p <= 1)

    def test_tukey_hsd(self):
        # Tukey HSD pairwise checks
        all_vals = self.g1 + self.g2 + self.g3
        all_labels = ["G1"]*5 + ["G2"]*5 + ["G3"]*5
        res = mc.pairwise_tukeyhsd(all_vals, all_labels, alpha=0.05)
        
        # Verify comparisons count
        self.assertEqual(len(res.meandiffs), 3)
        # Differences: G2-G1=1.0, G3-G1=2.0, G3-G2=1.0
        self.assertAlmostEqual(res.meandiffs[0], 1.0)
        self.assertAlmostEqual(res.meandiffs[1], 2.0)
        self.assertAlmostEqual(res.meandiffs[2], 1.0)

    def test_zero_variance_protection(self):
        # If a group has std == 0, we add a tiny epsilon to make variance positive
        zero_var_group = [2.0, 2.0, 2.0, 2.0, 2.0]
        
        # detect and apply protection
        if np.std(zero_var_group) == 0:
            zero_var_group = list(zero_var_group)
            zero_var_group[0] += 1e-6
            
        std_val = np.std(zero_var_group)
        self.assertTrue(std_val > 0)
        self.assertAlmostEqual(np.mean(zero_var_group), 2.0, places=5)

    def test_missing_values(self):
        # Drop NaNs without breaking alignment
        dirty_data = [1.2, None, 1.8, 1.4, np.nan, 1.6]
        clean_data = [float(x) for x in dirty_data if x is not None and not np.isnan(x)]
        
        self.assertEqual(len(clean_data), 4)
        self.assertEqual(clean_data, [1.2, 1.8, 1.4, 1.6])

    def test_effect_size_eta_squared(self):
        # We can calculate F-test and eta_squared manually for g1, g2, g3
        all_vals = self.g1 + self.g2 + self.g3
        grand_mean = np.mean(all_vals)
        ss_total = sum((x - grand_mean)**2 for x in all_vals)
        ss_between = len(self.g1)*(np.mean(self.g1) - grand_mean)**2 + \
                     len(self.g2)*(np.mean(self.g2) - grand_mean)**2 + \
                     len(self.g3)*(np.mean(self.g3) - grand_mean)**2
                     
        expected_eta = ss_between / ss_total
        
        # Now make an API call or check using the test client
        from app import app
        client = app.test_client()
        response = client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum')
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        
        # Verify eta_squared is present and between 0 and 1
        self.assertIn('eta_squared', data)
        self.assertIn('eta_interpretation', data)
        self.assertTrue(0 <= data['eta_squared'] <= 1)
        # Check interpretation category
        if data['eta_squared'] >= 0.14:
            self.assertEqual(data['eta_interpretation'], "Large")

        # Verify p-value display formatting fields exist
        self.assertIn('p_value_display', data['anova_table']['Between'])
        self.assertIn('p_value_display', data['shapiro_results'][0])
        self.assertIn('p_value_display', data['levene_result'])
        self.assertIn('p_adj_display', data['tukey_results'][0])

    def test_tukey_q_statistic(self):
        from app import app
        client = app.test_client()
        response = client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum')
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        
        # Verify that all tukey_results contain q_stat
        self.assertIn('tukey_results', data)
        for row in data['tukey_results']:
            self.assertIn('q_stat', row)
            self.assertTrue(row['q_stat'] >= 0)
            
            # Recalculate manually to verify Q = |mean1 - mean2| / SE_comp
            g1_name = row['group1']
            g2_name = row['group2']
            g1_vals = [pt['Value'] for pt in data['raw_data_points'] if pt['Group'] == g1_name]
            g2_vals = [pt['Value'] for pt in data['raw_data_points'] if pt['Group'] == g2_name]
            mean1 = np.mean(g1_vals)
            mean2 = np.mean(g2_vals)
            n1 = len(g1_vals)
            n2 = len(g2_vals)
            ms_within = data['anova_table']['Within']['MS']
            
            se_comparison = np.sqrt((ms_within / 2.0) * ((1.0 / n1) + (1.0 / n2)))
            expected_q = abs(mean1 - mean2) / se_comparison
            self.assertAlmostEqual(row['q_stat'], expected_q, places=2)

    def test_cld_letters_correctness(self):
        from app import app
        client = app.test_client()
        response = client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum')
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        
        # Verify tukey_letters are present
        self.assertIn('tukey_letters', data)
        letters = data['tukey_letters']
        
        # Group with the highest mean must have 'a'
        sorted_stats = sorted(data['summary_stats'], key=lambda s: s['Mean'], reverse=True)
        from app import format_group_name
        highest_group_formatted = format_group_name(sorted_stats[0]['Group'])
        self.assertIn(highest_group_formatted, letters)
        self.assertTrue(letters[highest_group_formatted].startswith('a'))
        
        # Letters must be lowercase strings containing alphabet letters
        for g, let in letters.items():
            self.assertTrue(let.islower())
            self.assertTrue(all('a' <= c <= 'z' for c in let))

    def test_control_response_and_null_state(self):
        from app import app
        client = app.test_client()
        
        # Case 1: Control group exists (Concentration factor grouped by Biochar)
        response1 = client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum')
        self.assertEqual(response1.status_code, 200)
        data1 = response1.get_json()
        self.assertIn('control_response', data1)
        self.assertIsNotNone(data1['control_response'])
        self.assertTrue(len(data1['control_response']) > 0)
        
        # Check that percent change is calculated correctly and interpretation matches
        for row in data1['control_response']:
            self.assertIn('treatment', row)
            self.assertIn('diff', row)
            self.assertIn('pct_change', row)
            self.assertIn('interpretation', row)
            
            # Verify interpretation based on pct_change
            pct = row['pct_change']
            interp = row['interpretation']
            if pct <= -10.0:
                self.assertEqual(interp, "Strong growth inhibition")
            elif pct < 0.0:
                self.assertEqual(interp, "Slight growth inhibition")
            elif pct < 10.0:
                self.assertEqual(interp, "Slight growth improvement")
            else:
                self.assertEqual(interp, "Substantial growth improvement")

        # Case 2: No Control group exists (e.g. Day factor grouped by Biochar and Concentration)
        response2 = client.get('/api/one-way?crop=Onion&variable=Root%20Length&factor=Day&biochar_filter=Acrostichum%20aureum&concentration_filter=0.5')
        self.assertEqual(response2.status_code, 200)
        data2 = response2.get_json()
        self.assertIn('control_response', data2)
        # Should be None/null because group names are Day 3, Day 5, Day 7 - no Control.
        self.assertIsNone(data2['control_response'])

class TestTwoWayANOVA(unittest.TestCase):
    def setUp(self):
        # Create an artificial unbalanced two-way dataset
        # Factor A: Biochar (B1, B2)
        # Factor B: Concentration (C0, C1, C2)
        # Replicated control (C0) has N=6, treatments have N=3
        self.data = []
        
        # B1 cells
        self.data.extend([{"Biochar": "B1", "Concentration": 0.0, "Value": v} for v in [1.0, 1.1, 0.9, 1.0, 1.2, 0.8]]) # C0
        self.data.extend([{"Biochar": "B1", "Concentration": 0.5, "Value": v} for v in [1.5, 1.6, 1.4]])             # C1
        self.data.extend([{"Biochar": "B1", "Concentration": 1.0, "Value": v} for v in [2.0, 2.2, 2.1]])             # C2

        # B2 cells
        self.data.extend([{"Biochar": "B2", "Concentration": 0.0, "Value": v} for v in [1.0, 1.1, 0.9, 1.0, 1.2, 0.8]]) # C0 (replicated)
        self.data.extend([{"Biochar": "B2", "Concentration": 0.5, "Value": v} for v in [1.2, 1.3, 1.1]])             # C1
        self.data.extend([{"Biochar": "B2", "Concentration": 1.0, "Value": v} for v in [1.4, 1.5, 1.3]])             # C2
        
        import pandas as pd
        self.df = pd.DataFrame(self.data)

    def test_twoway_anova_type3(self):
        import statsmodels.api as sm
        from statsmodels.formula.api import ols
        
        # Fit model with Sum contrast coding
        model = ols('Value ~ C(Biochar, Sum) * C(Concentration, Sum)', data=self.df).fit()
        anova_table = sm.stats.anova_lm(model, typ=3)
        
        # Total sample size N = 6 (B1 C0) + 3 (B1 C1) + 3 (B1 C2) + 6 (B2 C0) + 3 (B2 C1) + 3 (B2 C2) = 24
        # df_total = 23
        df_Biochar = anova_table.loc['C(Biochar, Sum)', 'df']
        df_Concentration = anova_table.loc['C(Concentration, Sum)', 'df']
        df_Interaction = anova_table.loc['C(Biochar, Sum):C(Concentration, Sum)', 'df']
        df_Residual = anova_table.loc['Residual', 'df']
        
        self.assertEqual(df_Biochar, 1)        # 2 biochars - 1
        self.assertEqual(df_Concentration, 2)  # 3 concentrations - 1
        self.assertEqual(df_Interaction, 2)    # 1 * 2
        self.assertEqual(df_Residual, 18)      # 24 - 6 cells = 18 degrees of freedom for error
        
        df_sum = df_Biochar + df_Concentration + df_Interaction + df_Residual
        self.assertEqual(df_sum, 23)

class TestAPIRoutes(unittest.TestCase):
    def setUp(self):
        from app import app
        self.client = app.test_client()
        self.client.testing = True

    def test_metadata_endpoint(self):
        response = self.client.get('/api/metadata')
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data['status'], 'success')
        self.assertIn('Onion', data['metadata'])
        self.assertIn('Pea', data['metadata'])

    def test_oneway_endpoint(self):
        response = self.client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum')
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data['status'], 'success')
        self.assertIn('summary_stats', data)
        self.assertIn('anova_table', data)
        self.assertIn('tukey_results', data)

    def test_twoway_endpoint(self):
        response = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207')
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data['status'], 'success')
        self.assertIn('anova_table', data)
        self.assertIn('cell_means', data)
        self.assertIn('posthoc_results', data)
        self.assertIn('interaction_plot_data', data)

    def test_export_excel_endpoint(self):
        # 1. Get stats from one-way endpoint
        oneway_resp = self.client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum')
        self.assertEqual(oneway_resp.status_code, 200)
        oneway_data = oneway_resp.get_json()

        # 2. Build payload with a minimal mock 1x1 pixel PNG in base64
        mock_png_base64 = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
        
        payload = {
            "crop": "Onion",
            "variable": "Root Length",
            "day": "Day 7",
            "factor": "Concentration",
            "biochar": "Acrostichum aureum",
            "summary_stats": oneway_data.get("summary_stats", []),
            "anova_table": oneway_data.get("anova_table", {}),
            "levene_result": oneway_data.get("levene_result", {}),
            "shapiro_results": oneway_data.get("shapiro_results", []),
            "tukey_results": oneway_data.get("tukey_results", []),
            "inference_summary": "Statistically Significant",
            "chart_image": mock_png_base64
        }

        # 3. Call export endpoint
        response = self.client.post('/api/export-excel', json=payload)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Verify it's a valid zip file structure (which Excel files are)
        import zipfile
        from io import BytesIO
        try:
            zip_file = zipfile.ZipFile(BytesIO(response.data))
            # Verify basic Excel internal files are present
            self.assertIn('[Content_Types].xml', zip_file.namelist())
            self.assertIn('xl/workbook.xml', zip_file.namelist())
        except Exception as zip_err:
            self.fail(f"Exported Excel file is not a valid zip archive: {str(zip_err)}")

    def test_configurable_alpha_oneway(self):
        # 1. Test default alpha
        resp_def = self.client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum')
        self.assertEqual(resp_def.status_code, 200)
        data_def = resp_def.get_json()
        self.assertEqual(data_def['alpha'], 0.05)

        # 2. Test valid alpha values
        for a in [0.001, 0.01, 0.05]:
            resp = self.client.get(f'/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum&alpha={a}')
            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertEqual(data['alpha'], a)
            
            # F-statistic and raw p-values must remain identical to default
            self.assertEqual(data['anova_table']['Between']['F'], data_def['anova_table']['Between']['F'])
            self.assertEqual(data['anova_table']['Between']['p_value'], data_def['anova_table']['Between']['p_value'])

        # 3. Test invalid alpha falls back to 0.05
        resp_inv = self.client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum&alpha=0.25')
        self.assertEqual(resp_inv.status_code, 200)
        data_inv = resp_inv.get_json()
        self.assertEqual(data_inv['alpha'], 0.05)

        # 4. Test that 0.10 is now rejected and falls back to 0.05
        resp_10 = self.client.get('/api/one-way?crop=Onion&variable=Root%20Length&day=Day%207&factor=Concentration&biochar_filter=Acrostichum%20aureum&alpha=0.10')
        self.assertEqual(resp_10.status_code, 200)
        data_10 = resp_10.get_json()
        self.assertEqual(data_10['alpha'], 0.05)

    def test_configurable_alpha_twoway(self):
        # 1. Test default alpha
        resp_def = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207')
        self.assertEqual(resp_def.status_code, 200)
        data_def = resp_def.get_json()
        self.assertEqual(data_def['alpha'], 0.05)

        # 2. Test valid alpha values
        for a in [0.001, 0.01, 0.05]:
            resp = self.client.get(f'/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&alpha={a}')
            self.assertEqual(resp.status_code, 200)
            data = resp.get_json()
            self.assertEqual(data['alpha'], a)
            
            # F-value and raw p-value must remain identical
            row_key = "C(Biochar, Sum):C(Concentration, Sum)"
            self.assertEqual(data['anova_table'][row_key]['F'], data_def['anova_table'][row_key]['F'])
            self.assertEqual(data['anova_table'][row_key]['p_value'], data_def['anova_table'][row_key]['p_value'])

        # 3. Test invalid alpha falls back to 0.05
        resp_inv = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&alpha=invalid')
        self.assertEqual(resp_inv.status_code, 200)
        data_inv = resp_inv.get_json()
        self.assertEqual(data_inv['alpha'], 0.05)

        # 4. Test that 0.10 is now rejected and falls back to 0.05
        resp_10 = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&alpha=0.10')
        self.assertEqual(resp_10.status_code, 200)
        data_10 = resp_10.get_json()
        self.assertEqual(data_10['alpha'], 0.05)

    def test_twoway_assumptions(self):
        # Query two-way API
        resp = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207')
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()

        # Check that assumptions fields are present
        self.assertIn('shapiro_results', data)
        self.assertIn('levene_result', data)

        shapiro = data['shapiro_results']
        levene = data['levene_result']

        # Assert correct keys in shapiro
        self.assertIn('statistic', shapiro)
        self.assertIn('p_value', shapiro)
        self.assertIn('normal', shapiro)
        self.assertIn('note', shapiro)

        # Assert correct keys in levene
        self.assertIn('statistic', levene)
        self.assertIn('p_value', levene)
        self.assertIn('equal_variance', levene)
        self.assertIn('note', levene)

        # Ensure statistics are floats or correct robust types
        if shapiro['statistic'] is not None:
            self.assertIsInstance(shapiro['statistic'], float)
            self.assertIsInstance(shapiro['p_value'], float)
            self.assertIsInstance(shapiro['normal'], bool)
        
        if levene['statistic'] is not None:
            self.assertIsInstance(levene['statistic'], float)
            self.assertIsInstance(levene['p_value'], float)
            self.assertIsInstance(levene['equal_variance'], bool)

class TestTwoWaySubsetSelection(unittest.TestCase):
    def setUp(self):
        from app import app
        self.client = app.test_client()
        self.client.testing = True

    def test_twoway_full_set_identity(self):
        # 1. Fetch baseline (without biochars parameter, i.e., all biochars selected)
        resp_baseline = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207')
        self.assertEqual(resp_baseline.status_code, 200)
        data_baseline = resp_baseline.get_json()

        # 2. Fetch with all 4 biochars explicitly in parameter
        all_biochars = "Acrostichum aureum,Cyclosorus interruptus,Ludwigia peruviana,Quisqualis indica"
        resp_all = self.client.get(f'/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&biochars={all_biochars}')
        self.assertEqual(resp_all.status_code, 200)
        data_all = resp_all.get_json()

        # 3. Assert exact statistical identities in the ANOVA table
        baseline_anova = data_baseline['anova_table']
        all_anova = data_all['anova_table']
        
        for key in baseline_anova.keys():
            self.assertEqual(baseline_anova[key]['SS'], all_anova[key]['SS'])
            self.assertEqual(baseline_anova[key]['df'], all_anova[key]['df'])
            self.assertEqual(baseline_anova[key]['MS'], all_anova[key]['MS'])
            self.assertEqual(baseline_anova[key]['F'], all_anova[key]['F'])
            self.assertEqual(baseline_anova[key]['p_value'], all_anova[key]['p_value'])

    def test_twoway_subset_filtering(self):
        # Select subset: only Acrostichum aureum and Quisqualis indica
        subset_biochars = "Acrostichum aureum,Quisqualis indica"
        resp = self.client.get(f'/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&biochars={subset_biochars}')
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()

        # Verify that cell_means and posthoc only include the selected biochars
        cell_means_biochars = [row['Biochar'] for row in data['cell_means']]
        self.assertEqual(sorted(cell_means_biochars), ["Acrostichum aureum", "Quisqualis indica"])

        posthoc_biochars = list(data['posthoc_results'].keys())
        self.assertEqual(sorted(posthoc_biochars), ["Acrostichum aureum", "Quisqualis indica"])

        df_ver = data['debug_details']['df_verification']
        self.assertTrue(df_ver['df_verified'])

    def test_twoway_subset_validation_error(self):
        # Query with only 1 biochar selected
        resp = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&biochars=Acrostichum%20aureum')
        self.assertEqual(resp.status_code, 400)
        data = resp.get_json()
        self.assertEqual(data['status'], 'error')
        self.assertIn("minimum 2 required", data['message'])

    def test_twoway_control_mode_replicated_default(self):
        # Default query
        resp_def = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207')
        self.assertEqual(resp_def.status_code, 200)
        data_def = resp_def.get_json()
        
        # Explicit replicated query
        resp_repl = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&control_mode=replicated')
        self.assertEqual(resp_repl.status_code, 200)
        data_repl = resp_repl.get_json()
        
        # Assert equal results
        self.assertEqual(data_def['control_mode'], 'replicated')
        self.assertEqual(data_repl['control_mode'], 'replicated')
        self.assertEqual(data_def['anova_table']['Residual']['SS'], data_repl['anova_table']['Residual']['SS'])

    def test_twoway_control_mode_exclude(self):
        # Exclude shared control query
        resp = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&control_mode=exclude')
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()
        
        self.assertEqual(data['control_mode'], 'exclude')
        self.assertEqual(data['control_mode_label'], 'Exclude Shared Control')
        
        # Concentration 0.0 must be completely absent from factor levels
        concs = data['debug_details']['factor_levels']['Concentration']
        self.assertNotIn(0.0, concs)
        self.assertNotIn("0.0", concs)
        self.assertNotIn(0, concs)
        
        # Cell means grid must not have control
        for row in data['cell_means']:
            self.assertNotIn("0.0", row)
            self.assertNotIn("0", row)
            
        # Tukey table must not contain Control vs ...
        for b_key, comparisons in data['posthoc_results'].items():
            for comp in comparisons:
                self.assertNotEqual(comp['group1'], 'Control')
                self.assertNotEqual(comp['group2'], 'Control')
                
        # Verification that degree of freedom of Concentration corresponds to 4 levels (0.1, 0.2, 0.3, 0.5 -> df = 3)
        df_conc = data['anova_table']['C(Concentration, Sum)']['df']
        self.assertEqual(df_conc, 3)

    def test_twoway_bidirectional_sme(self):
        # 1. Default replicated query
        resp = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207')
        self.assertEqual(resp.status_code, 200)
        data = resp.get_json()

        self.assertIn('simple_main_effects', data)
        sme = data['simple_main_effects']
        self.assertIn('within_biochar', sme)
        self.assertIn('within_concentration', sme)

        # Check within_biochar metadata
        wb = sme['within_biochar']
        self.assertEqual(wb['selector_type'], 'biochar')
        self.assertIsInstance(wb['selector_values'], list)
        self.assertIn('Acrostichum aureum', wb['selector_values'])
        self.assertIn('Acrostichum aureum', wb['results'])

        # Check within_concentration metadata
        wc = sme['within_concentration']
        self.assertEqual(wc['selector_type'], 'concentration')
        self.assertIsInstance(wc['selector_values'], list)
        # Replicated mode: Control should be present in concentration values
        self.assertIn('Control', wc['selector_values'])
        self.assertIn('Control', wc['results'])

        # Verify comparison fields are present and preformatted correctly
        for b_val in wb['results']['Acrostichum aureum']:
            self.assertIn('comparison', b_val)
            self.assertIn(' vs ', b_val['comparison'])
            self.assertEqual(b_val['comparison'], f"{b_val['group1']} vs {b_val['group2']}")

        for c_val in wc['results']['Control']:
            self.assertIn('comparison', c_val)
            self.assertIn(' vs ', c_val['comparison'])
            self.assertEqual(c_val['comparison'], f"{c_val['group1']} vs {c_val['group2']}")

        # Verify legacy posthoc_results is identical to simple_main_effects within_biochar
        self.assertEqual(data['posthoc_results'], wb['results'])

        # 2. Exclude shared control query
        resp_ex = self.client.get('/api/two-way?crop=Onion&variable=Root%20Length&day=Day%207&control_mode=exclude')
        self.assertEqual(resp_ex.status_code, 200)
        data_ex = resp_ex.get_json()

        sme_ex = data_ex['simple_main_effects']
        wc_ex = sme_ex['within_concentration']
        
        # Exclude mode: Control should be absent in concentration values
        self.assertNotIn('Control', wc_ex['selector_values'])
        self.assertNotIn('Control', wc_ex['results'])

    def test_twoway_excel_export_structure(self):
        import io
        import openpyxl
        from app import run_two_way_analysis
        
        # 1. Export Excel using the new endpoint
        payload = {
            "crop": "Onion",
            "variable": "Root Length",
            "day": "Day 7",
            "biochars": "Acrostichum aureum,Quisqualis indica",
            "control_mode": "replicated",
            "alpha": 0.05
        }
        response = self.client.post('/api/export-excel-twoway', json=payload)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # 2. Load workbook
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        
        # 3. Verify worksheet names
        expected_sheets = [
            "Analysis Summary",
            "Experimental Design",
            "Cell Means Matrix",
            "Type III ANOVA",
            "Assumption Diagnostics",
            "Simple Main Effects",
            "Interaction Plot Data",
            "Analysis Dataset"
        ]
        self.assertEqual(wb.sheetnames, expected_sheets)
        
        # 4. Verify representative cell values in Sheet 1: Analysis Summary
        ws1 = wb["Analysis Summary"]
        self.assertEqual(ws1["A1"].value, "Two-Way ANOVA Analysis Summary")
        self.assertEqual(ws1["A4"].value, "Crop")
        self.assertEqual(ws1["B4"].value, "Onion")
        self.assertEqual(ws1["A5"].value, "Variable")
        self.assertEqual(ws1["B5"].value, "Root Length")
        self.assertEqual(ws1["A6"].value, "Day")
        self.assertEqual(ws1["B6"].value, "Day 7")
        self.assertEqual(ws1["A11"].value, "Control Handling Mode")
        self.assertEqual(ws1["B11"].value, "Include Independent Controls (Default)")
        
        # 5. Verify Sheet 2: Experimental Design
        ws2 = wb["Experimental Design"]
        self.assertEqual(ws2["A1"].value, "Experimental Design Summary")
        self.assertEqual(ws2["A6"].value, "Active Biochar Species")
        self.assertEqual(ws2["B6"].value, "Acrostichum aureum, Quisqualis indica")
        self.assertEqual(ws2["A7"].value, "Active Concentration Levels")
        self.assertEqual(ws2["B7"].value, "Control, 0.1 g/L, 0.2 g/L, 0.3 g/L, 0.5 g/L")
        
        # 6. Verify Sheet 3: Cell Means Matrix
        ws3 = wb["Cell Means Matrix"]
        self.assertEqual(ws3["A1"].value, "Cell Replication & Means Matrix (Biochar Species × Concentration)")
        self.assertEqual(ws3["A3"].value, "Biochar Species")
        self.assertEqual(ws3["B3"].value, "Control")
        self.assertEqual(ws3["C3"].value, "0.1 g/L")
        self.assertEqual(ws3["A4"].value, "Acrostichum aureum")
        self.assertEqual(ws3["A5"].value, "Quisqualis indica")
        
        val_cell = ws3["B4"].value
        self.assertIn("SD: ", val_cell)
        self.assertIn("(N=", val_cell)
        
        # 7. Verify Sheet 4: Type III ANOVA
        ws4 = wb["Type III ANOVA"]
        self.assertEqual(ws4["A1"].value, "Type III ANOVA Table (For Unbalanced Designs)")
        self.assertEqual(ws4["A5"].value, "Biochar Species (Factor A)")
        self.assertEqual(ws4["A6"].value, "Concentration (Factor B)")
        self.assertEqual(ws4["A7"].value, "Biochar Species & Concentration (Interaction)")
        self.assertEqual(ws4["A8"].value, "Error (Residuals)")
        self.assertEqual(ws4["A9"].value, "Total (Corrected)")
        
        self.assertEqual(ws4["C5"].value, 1) # 2 biochars -> 1 df
        self.assertEqual(ws4["C6"].value, 4) # 5 concentrations -> 4 df
        self.assertEqual(ws4["C7"].value, 4) # Interaction -> 4 df
        
        err_df = ws4["C8"].value
        total_df = ws4["C9"].value
        self.assertEqual(total_df, 1 + 4 + 4 + err_df)
        
        # 8. Verify Sheet 5: Assumption Diagnostics
        ws5 = wb["Assumption Diagnostics"]
        self.assertEqual(ws5["A1"].value, "Homogeneity of Variance (Levene's Test)")
        self.assertEqual(ws5["A11"].value, "Normality of Residuals (Shapiro-Wilk Test)")
        
        # 9. Verify Sheet 6: Simple Main Effects
        ws6 = wb["Simple Main Effects"]
        self.assertEqual(ws6["A1"].value, "Post-hoc Analysis of Simple Main Effects (Tukey HSD)")
        col_A_vals = [cell.value for cell in ws6["A"]]
        self.assertIn("Section B: Compare Biochar Species within Concentration", col_A_vals)
        
        # 10. Verify Sheet 8: Analysis Dataset integrity
        ws8 = wb["Analysis Dataset"]
        self.assertEqual(ws8["A1"].value, "Analysis Dataset (OLS Model Inputs)")
        
        # Verify dataset matches df used for model fitting
        res_dict, combined_df = run_two_way_analysis(
            "Onion",
            "Root Length",
            "Day 7",
            "Acrostichum aureum,Quisqualis indica",
            "replicated",
            "0.05"
        )
        sheet_rows = []
        for r_idx in range(4, ws8.max_row + 1):
            b_val = ws8.cell(row=r_idx, column=1).value
            c_val = ws8.cell(row=r_idx, column=2).value
            v_val = ws8.cell(row=r_idx, column=3).value
            if b_val is not None:
                sheet_rows.append({
                    "Biochar": b_val,
                    "Concentration": float(c_val),
                    "Value": float(v_val)
                })
        
        df_rows = []
        for idx, r in combined_df.iterrows():
            df_rows.append({
                "Biochar": r["Biochar"],
                "Concentration": float(r["Concentration"]),
                "Value": float(r["Value"])
            })
            
        self.assertEqual(len(sheet_rows), len(df_rows))
        for s_row, d_row in zip(sheet_rows, df_rows):
            self.assertEqual(s_row["Biochar"], d_row["Biochar"])
            self.assertAlmostEqual(s_row["Concentration"], d_row["Concentration"])
            self.assertAlmostEqual(s_row["Value"], d_row["Value"], places=4)

    def test_twoway_excel_export_exclude_control(self):
        import io
        import openpyxl
        from app import run_two_way_analysis
        
        payload = {
            "crop": "Onion",
            "variable": "Root Length",
            "day": "Day 7",
            "biochars": "Acrostichum aureum,Quisqualis indica",
            "control_mode": "exclude",
            "alpha": 0.05
        }
        response = self.client.post('/api/export-excel-twoway', json=payload)
        self.assertEqual(response.status_code, 200)
        
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        
        # Verify Sheet 1 control mode value
        ws1 = wb["Analysis Summary"]
        self.assertEqual(ws1["B11"].value, "Exclude Shared Control")
        
        # Verify Sheet 2 active concentration levels does not contain "Control"
        ws2 = wb["Experimental Design"]
        self.assertEqual(ws2["B7"].value, "0.1 g/L, 0.2 g/L, 0.3 g/L, 0.5 g/L")
        
        # Verify Sheet 3 cell means matrix columns do not contain "Control"
        ws3 = wb["Cell Means Matrix"]
        headers = [cell.value for cell in ws3[3] if cell.value is not None]
        self.assertNotIn("Control", headers)
        
        # Verify Sheet 4: df of Concentration (Factor B) is 3 (levels: 0.1, 0.2, 0.3, 0.5 -> 3 df)
        ws4 = wb["Type III ANOVA"]
        self.assertEqual(ws4["C6"].value, 3)
        self.assertEqual(ws4["C7"].value, 3)
        
        # Verify Sheet 6: SME comparisons do not contain "Control"
        ws6 = wb["Simple Main Effects"]
        for row in ws6.iter_rows(values_only=True):
            if row[0] is not None and "Section A" in str(row[0]):
                continue
            if row[0] is not None and "Section B" in str(row[0]):
                continue
            if row[1] is not None:
                self.assertNotIn("Control", str(row[1]))
                self.assertNotIn("0.0", str(row[1]))
            if row[0] is not None:
                self.assertNotIn("Control", str(row[0]))
                
        # Verify Sheet 8: Analysis Dataset does not contain Concentration = 0.0, and matches model fitting dataframe
        ws8 = wb["Analysis Dataset"]
        for r_idx in range(4, ws8.max_row + 1):
            conc_val = ws8.cell(row=r_idx, column=2).value
            if conc_val is not None:
                self.assertNotEqual(float(conc_val), 0.0)

        # Verify dataset matches df used for model fitting
        res_dict, combined_df = run_two_way_analysis(
            "Onion",
            "Root Length",
            "Day 7",
            "Acrostichum aureum,Quisqualis indica",
            "exclude",
            "0.05"
        )
        sheet_rows = []
        for r_idx in range(4, ws8.max_row + 1):
            b_val = ws8.cell(row=r_idx, column=1).value
            c_val = ws8.cell(row=r_idx, column=2).value
            v_val = ws8.cell(row=r_idx, column=3).value
            if b_val is not None:
                sheet_rows.append({
                    "Biochar": b_val,
                    "Concentration": float(c_val),
                    "Value": float(v_val)
                })
        
        df_rows = []
        for idx, r in combined_df.iterrows():
            df_rows.append({
                "Biochar": r["Biochar"],
                "Concentration": float(r["Concentration"]),
                "Value": float(r["Value"])
            })
            
        self.assertEqual(len(sheet_rows), len(df_rows))
        for s_row, d_row in zip(sheet_rows, df_rows):
            self.assertEqual(s_row["Biochar"], d_row["Biochar"])
            self.assertAlmostEqual(s_row["Concentration"], d_row["Concentration"])
            self.assertAlmostEqual(s_row["Value"], d_row["Value"], places=4)


if __name__ == "__main__":
    unittest.main()


