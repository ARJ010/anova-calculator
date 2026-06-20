import os
import io
import base64
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from flask import Flask, jsonify, render_template, request, send_file

app = Flask(__name__)

# Resolve workbook path relative to the app file location
BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = BASE_DIR / "Base" / "Biochar_Onion_Pea_Final.xlsx"

# Master flat dataframe cache
FLAT_DF = None

def clean_biochar_name(sheet_name):
    name = sheet_name.strip().lower()
    if "acrostichum" in name:
        return "Acrostichum aureum"
    elif "cyclosorus" in name:
        return "Cyclosorus interruptus"
    elif "ludwigia" in name or "ludwegia" in name:
        return "Ludwigia peruviana"
    elif "quisqualis" in name:
        return "Quisqualis indica"
    elif "control" in name:
        return "Control"
    return sheet_name.strip()

def is_numeric(val):
    if val is None:
        return False
    try:
        float(val)
        return True
    except ValueError:
        return False

def flatten_default_workbook():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found at {WORKBOOK_PATH}")

    # Load workbook with data_only=True to get computed values, not formulas
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    
    rows_data = []

    # Map sheets exactly
    onion_sheets = ['Control Onion', 'Acrostichum aureum', 'Cyclosorus interruptus', 'Ludwigia peruviana', 'Quisqualis indica']
    pea_shoot_sheets = ['Control Pea Shoot', 'Acrostichum shoot ', 'Quisqualis indica Shoot', 'Ludwigia peruviana Shoot', 'Cyclosorus interruptus Shoot']
    pea_root_sheets = ['Control Pea Root', 'Acrostichum root (Pea)', 'Quisqualis root (Pea)', 'Ludwegia root (Pea) ', 'Cyclosorus root (Pea) ']
    pea_tl_sheets = ['Control Pea Total Length', 'Cyclosorus TL (Pea)  ', 'Ludwegia TL(Pea) ', 'Quisqualis TL (Pea)', 'Acrostichum TL (Pea) ']

    all_sheets = onion_sheets + pea_shoot_sheets + pea_root_sheets + pea_tl_sheets

    for sheet_name in all_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet {sheet_name} not found in workbook.")
            continue

        ws = wb[sheet_name]
        biochar = clean_biochar_name(sheet_name)
        
        # Determine Crop and Variable
        if sheet_name in onion_sheets:
            crop = "Onion"
            variable = "Root Length"
        else:
            crop = "Pea"
            if sheet_name in pea_shoot_sheets:
                variable = "Shoot Length"
            elif sheet_name in pea_root_sheets:
                variable = "Root Length"
            else:
                variable = "Total Length"

        # Determine Replicate Columns range
        # Control Onion has 20 replicates (Cols C to V -> index 2 to 21)
        # Control Pea sheets have 10 replicates (Cols C to L -> index 2 to 11)
        # Treatment Onion has 10 replicates (Cols C to L -> index 2 to 11)
        # Treatment Pea has 5 replicates (Cols C to G -> index 2 to 6)
        is_control = (biochar == "Control")
        if is_control:
            if crop == "Onion":
                col_start, col_end = 2, 21  # C to V
            else:
                col_start, col_end = 2, 11  # C to L
        else:
            if crop == "Onion":
                col_start, col_end = 2, 11  # C to L
            else:
                col_start, col_end = 2, 6   # C to G

        # Determine row processing limit to exclude summary blocks at the bottom of the sheets
        if is_control:
            if crop == "Onion" or sheet_name == "Control Pea Shoot":
                max_rows = 4  # Rows 2-4 (indices 1, 2, 3)
            else:
                max_rows = 2  # Row 2 (index 1)
        else:
            if crop == "Onion" or sheet_name in pea_shoot_sheets:
                max_rows = 17 # Rows 2-5, 8-11, 14-17 (indices up to 16)
            else:
                max_rows = 5  # Rows 2-5 (indices 1, 2, 3, 4)

        # Parse rows
        current_day = None
        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if r_idx >= max_rows:
                break

            # Check for Day headers in Column A
            val_a = row[0]
            if val_a is not None and str(val_a).strip().startswith("Day"):
                current_day = str(val_a).strip()
                if not is_control:
                    continue  # Treatment day headers don't contain data rows
            
            # For Pea Root and Pea Total Length sheets, there is no Day header, it's always Day 7
            if current_day is None and (sheet_name in pea_root_sheets or sheet_name in pea_tl_sheets):
                current_day = "Day 7"

            # Check if row is a data row
            val_b = row[1]
            
            # Determine Concentration
            conc = None
            if is_control:
                # Control rows have Concentration = 0.0, and are identified by the Day label in Column A
                if val_a is not None and str(val_a).strip() == current_day:
                    conc = 0.0
            else:
                # Treatment rows have the concentration value in Column B
                if val_b is not None and is_numeric(val_b):
                    conc = float(val_b)

            if conc is not None and current_day is not None:
                # Extract replicate values
                for col_idx in range(col_start, col_end + 1):
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val is not None and is_numeric(val):
                            rows_data.append({
                                "Crop": crop,
                                "Biochar": biochar,
                                "Concentration": conc,
                                "Day": current_day,
                                "Variable": variable,
                                "Value": float(val)
                            })

    df = pd.DataFrame(rows_data)
    return df

# Initialize dataframe at startup
try:
    FLAT_DF = flatten_default_workbook()
    print(f"Data pipeline successfully initialized. Loaded {len(FLAT_DF)} rows.")
except Exception as e:
    print(f"CRITICAL: Failed to load default workbook: {str(e)}")

@app.route("/debug")
def debug_data():
    if FLAT_DF is None:
        return jsonify({"status": "error", "message": "DataFrame not loaded"}), 500
    
    # Calculate unique counts and basic descriptions
    summary = {
        "status": "success",
        "shape": FLAT_DF.shape,
        "unique_crops": FLAT_DF["Crop"].unique().tolist(),
        "unique_biochars": FLAT_DF["Biochar"].unique().tolist(),
        "unique_concentrations": FLAT_DF["Concentration"].unique().tolist(),
        "unique_days": FLAT_DF["Day"].unique().tolist(),
        "unique_variables": FLAT_DF["Variable"].unique().tolist(),
        "missing_values": FLAT_DF.isnull().sum().to_dict(),
        "sample_counts_by_crop": FLAT_DF.groupby("Crop").size().to_dict(),
        "sample_counts_by_variable": FLAT_DF.groupby("Variable").size().to_dict(),
        "sample_counts_by_biochar": FLAT_DF.groupby("Biochar").size().to_dict(),
        "group_summary_preview": FLAT_DF.groupby(["Crop", "Biochar", "Day", "Variable", "Concentration"]).agg(
            Count=("Value", "count"),
            Mean=("Value", "mean"),
            Variance=("Value", "var")
        ).reset_index().head(30).to_dict(orient="records"),
        "preview_rows": FLAT_DF.head(20).to_dict(orient="records")
    }
    return jsonify(summary)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/metadata")
def api_metadata():
    if FLAT_DF is None:
        return jsonify({"status": "error", "message": "DataFrame not loaded"}), 500
        
    crops = FLAT_DF["Crop"].unique().tolist()
    
    metadata = {}
    for crop in crops:
        crop_df = FLAT_DF[FLAT_DF["Crop"] == crop]
        metadata[crop] = {
            "variables": crop_df["Variable"].unique().tolist(),
            "biochars": sorted([b for b in crop_df["Biochar"].unique().tolist() if b != "Control"]),
            "days": sorted(crop_df["Day"].unique().tolist()),
            "concentrations": sorted([float(c) for c in crop_df["Concentration"].unique().tolist() if c > 0.0])
        }
        
    return jsonify({
        "status": "success",
        "metadata": metadata
    })

def format_group_name(name):
    name_str = str(name).strip().lower()
    if name_str in ["control", "0.0", "0"]:
        return "Control"
    return str(name).strip()

def format_p_value(p):
    """
    Publication-safe p-value formatter.
    """
    if p is None:
        return "N/A"
    try:
        p_val = float(p)
    except (ValueError, TypeError):
        return str(p)
        
    if p_val == 0.0:
        return "< 0.0001 (<1e-16)"
        
    if p_val >= 0.0001:
        return f"{p_val:.4f}"
    else:
        formatted_sci = f"{p_val:.2e}"
        if formatted_sci.startswith("0.00") or formatted_sci == "0.00e+00":
            return "< 0.0001 (<1e-16)"
        return f"< 0.0001 ({formatted_sci})"

def validate_alpha(alpha_param):
    """
    Validates alpha value. Allowed values are 0.001, 0.01, 0.05.
    Silently falls back to 0.05 on any invalid value.
    """
    try:
        alpha = float(alpha_param)
    except (TypeError, ValueError):
        return 0.05
    allowed_alphas = [0.001, 0.01, 0.05]
    for a in allowed_alphas:
        if abs(alpha - a) < 1e-6:
            return a
    return 0.05

def get_compact_letter_display(groups, tukey_results, group_means):
    """
    groups: list of group names (strings)
    tukey_results: list of dicts, each with group1, group2, reject (bool)
    group_means: dict of {group_name: mean_value}
    """
    n_groups = len(groups)
    if n_groups == 0:
        return {}
    
    # Sort groups descending by mean to rank them
    sorted_groups_by_mean = sorted(groups, key=lambda g: group_means.get(g, 0.0), reverse=True)
    group_ranks = {g: idx for idx, g in enumerate(sorted_groups_by_mean)}
    
    # Build adjacency list of non-significance graph (edges exist if reject is False)
    adj = {g: set() for g in groups}
    
    significant_pairs = set()
    for res in tukey_results:
        g1, g2 = res['group1'], res['group2']
        if res['reject']:
            significant_pairs.add((g1, g2))
            significant_pairs.add((g2, g1))
            
    for i in range(len(groups)):
        for j in range(i + 1, len(groups)):
            g1, g2 = groups[i], groups[j]
            if (g1, g2) not in significant_pairs:
                adj[g1].add(g2)
                adj[g2].add(g1)
                
    # Find all maximal cliques using Bron-Kerbosch algorithm
    cliques = []
    def bron_kerbosch(r, p, x):
        if not p and not x:
            cliques.append(r)
            return
        # Deterministic order sorted by mean rank
        for vertex in sorted(list(p), key=lambda g: group_ranks.get(g, 999)):
            bron_kerbosch(
                r.union([vertex]),
                p.intersection(adj[vertex]),
                x.intersection(adj[vertex])
            )
            p.remove(vertex)
            x.add(vertex)
            
    bron_kerbosch(set(), set(groups), set())
    
    # Sort cliques deterministically by the sorted list of their group ranks
    cliques.sort(key=lambda c: sorted([group_ranks[g] for g in c]))
    
    # Assign lowercase letters ('a', 'b', 'c', ...) to cliques
    group_letters = {g: [] for g in groups}
    for idx, clique in enumerate(cliques):
        letter = chr(97 + idx) # 0 -> 'a', 1 -> 'b', etc.
        for g in clique:
            group_letters[g].append(letter)
            
    # Concatenate letters and sort alphabetically (e.g. 'ab')
    # Unify key naming: map format_group_name keys to the output dict
    return {format_group_name(g): "".join(sorted(letters)) for g, letters in group_letters.items()}

@app.route("/api/one-way", methods=["POST", "GET"])
def api_one_way():
    if FLAT_DF is None:
        return jsonify({"status": "error", "message": "Data pipeline not initialized"}), 500

    data = request.args if request.method == "GET" else request.get_json(silent=True)
    if not data:
        data = request.form

    crop = data.get("crop")
    variable = data.get("variable")
    day = data.get("day")
    factor = data.get("factor")
    biochar_filter = data.get("biochar_filter")
    
    conc_filter_val = data.get("concentration_filter")
    concentration_filter = float(conc_filter_val) if conc_filter_val is not None and is_numeric(conc_filter_val) else None

    # Parse and validate alpha (nominal significance level)
    alpha = validate_alpha(data.get("alpha", "0.05"))

    # Filter master dataframe
    df_filtered = FLAT_DF.copy()
    df_filtered = df_filtered[df_filtered["Crop"] == crop]
    df_filtered = df_filtered[df_filtered["Variable"] == variable]

    groups_dict = {}
    formula_used = ""
    excluded_groups = []
    missing_dropped = 0

    if factor == "Concentration":
        df_filtered = df_filtered[df_filtered["Day"] == day]
        df_filtered = df_filtered[
            (df_filtered["Biochar"] == biochar_filter) | (df_filtered["Biochar"] == "Control")
        ]
        
        groups = df_filtered.groupby("Concentration")
        formula_used = f"Value ~ Concentration (grouped by Biochar: {biochar_filter}, Day: {day})"
        
        for name, group in groups:
            groups_dict[str(name)] = group["Value"].dropna().tolist()

    elif factor == "Biochar":
        df_filtered = df_filtered[df_filtered["Day"] == day]
        df_filtered = df_filtered[df_filtered["Concentration"] == concentration_filter]
        
        groups = df_filtered.groupby("Biochar")
        formula_used = f"Value ~ Biochar (grouped by Concentration: {concentration_filter}, Day: {day})"
        
        for name, group in groups:
            groups_dict[str(name)] = group["Value"].dropna().tolist()

    elif factor == "Day":
        if biochar_filter == "Control":
            df_filtered = df_filtered[(df_filtered["Biochar"] == "Control") & (df_filtered["Concentration"] == 0.0)]
        else:
            df_filtered = df_filtered[(df_filtered["Biochar"] == biochar_filter) & (df_filtered["Concentration"] == concentration_filter)]
            
        groups = df_filtered.groupby("Day")
        formula_used = f"Value ~ Day (grouped by Biochar: {biochar_filter}, Concentration: {concentration_filter})"
        
        for name, group in groups:
            groups_dict[str(name)] = group["Value"].dropna().tolist()

    else:
        return jsonify({"status": "error", "message": "Invalid factor specified"}), 400

    # Zero variance protection & NaN cleaning
    sanitized_groups = {}
    for g_name, g_vals in list(groups_dict.items()):
        g_vals = [float(x) for x in g_vals if x is not None and is_numeric(x)]
        if len(g_vals) < 2:
            excluded_groups.append(f"{g_name} (sample size < 2)")
            continue
            
        # Detect zero variance
        import numpy as np
        if np.std(g_vals) == 0:
            g_vals = list(g_vals)
            g_vals[0] += 1e-6 # add epsilon to first element
            
        sanitized_groups[g_name] = g_vals

    if len(sanitized_groups) < 2:
        return jsonify({
            "status": "error",
            "message": f"Not enough valid groups to perform ANOVA. Discovered groups: {list(groups_dict.keys())}. Valid groups: {list(sanitized_groups.keys())}."
        }), 400

    import numpy as np
    import scipy.stats as stats
    import statsmodels.stats.multicomp as mc

    # 1. Summary Statistics for each group (Astatsa-style)
    summary_stats = []
    def sorting_key(group_name):
        name_str = str(group_name).strip()
        if name_str.lower() in ["control", "0.0", "0"]:
            return (0, 0.0, name_str)
        try:
            val = float(name_str)
            return (1, val, name_str)
        except ValueError:
            return (2, name_str.lower(), name_str)

    group_names = sorted(list(sanitized_groups.keys()), key=sorting_key)
    
    anova_inputs = [sanitized_groups[name] for name in group_names]

    all_values = []
    all_group_labels = []
    
    for name in group_names:
        vals = sanitized_groups[name]
        n = len(vals)
        sum_x = sum(vals)
        mean = sum_x / n
        var = np.var(vals, ddof=1)
        sd = np.sqrt(var)
        se = sd / np.sqrt(n) if n > 0 else 0.0
        sum_x2 = sum(x**2 for x in vals)
        
        summary_stats.append({
            "Group": name,
            "N": n,
            "Sum": round(sum_x, 4),
            "Mean": round(mean, 4),
            "Variance": round(var, 4),
            "SD": round(sd, 4),
            "SE": round(float(se), 4),
            "SumSq": round(sum_x2, 4)
        })
        
        all_values.extend(vals)
        all_group_labels.extend([name] * n)

    # 2. Assumptions: Shapiro-Wilk (Normality) for each group
    shapiro_results = []
    for name in group_names:
        vals = sanitized_groups[name]
        if len(vals) >= 3:
            sh_stat, sh_p = stats.shapiro(vals)
            shapiro_results.append({
                "Group": name,
                "Statistic": round(sh_stat, 4),
                "p_value": round(sh_p, 4),
                "p_value_display": format_p_value(sh_p),
                "Normal": bool(sh_p >= alpha)
            })
        else:
            shapiro_results.append({
                "Group": name,
                "Statistic": None,
                "p_value": None,
                "p_value_display": "N/A",
                "Normal": None,
                "Note": "N < 3"
            })

    # 3. Assumptions: Levene's Test (Homogeneity of Variance)
    lev_stat, lev_p = stats.levene(*anova_inputs)
    levene_result = {
        "Statistic": round(lev_stat, 4),
        "p_value": round(lev_p, 4),
        "p_value_display": format_p_value(lev_p),
        "Equal_Variance": bool(lev_p >= alpha)
    }

    # 4. Standard One-Way ANOVA
    f_stat, f_p = stats.f_oneway(*anova_inputs)
    
    k = len(group_names)
    N = len(all_values)
    df_between = k - 1
    df_within = N - k
    df_total = N - 1
    
    grand_mean = sum(all_values) / N
    ss_total = sum((x - grand_mean)**2 for x in all_values)
    ss_between = sum(len(sanitized_groups[name]) * (np.mean(sanitized_groups[name]) - grand_mean)**2 for name in group_names)
    ss_within = ss_total - ss_between
    
    ms_between = ss_between / df_between
    ms_within = ss_within / df_within
    
    anova_table = {
        "Between": {
            "SS": round(ss_between, 4),
            "df": df_between,
            "MS": round(ms_between, 4),
            "F": round(f_stat, 4),
            "p_value": round(f_p, 4),
            "p_value_display": format_p_value(f_p)
        },
        "Within": {
            "SS": round(ss_within, 4),
            "df": df_within,
            "MS": round(ms_within, 4)
        },
        "Total": {
            "SS": round(ss_total, 4),
            "df": df_total
        },
        "Significant": bool(f_p < alpha)
    }

    # Calculate Effect Size (eta squared)
    eta_squared = ss_between / ss_total if ss_total > 0 else 0.0
    if eta_squared < 0.01:
        eta_interpretation = "Negligible"
    elif eta_squared < 0.06:
        eta_interpretation = "Small"
    elif eta_squared < 0.14:
        eta_interpretation = "Moderate"
    else:
        eta_interpretation = "Large"

    # 5. Tukey HSD Post-hoc Test
    tukey_res = mc.pairwise_tukeyhsd(np.array(all_values), np.array(all_group_labels), alpha=alpha)
    tukey_table = []
    
    for row in tukey_res._results_table.data[1:]:
        g1, g2, meandiff, p_adj, lower, upper, reject = row
        
        # Calculate Tukey Q statistic
        n1 = len(sanitized_groups[str(g1)])
        n2 = len(sanitized_groups[str(g2)])
        mean1 = np.mean(sanitized_groups[str(g1)])
        mean2 = np.mean(sanitized_groups[str(g2)])
        se_comparison = np.sqrt((ms_within / 2.0) * ((1.0 / n1) + (1.0 / n2)))
        q_stat = abs(mean1 - mean2) / se_comparison if se_comparison > 0 else 0.0
        
        tukey_table.append({
            "group1": str(g1),
            "group2": str(g2),
            "meandiff": round(float(meandiff), 4),
            "p_adj": round(float(p_adj), 4) if isinstance(p_adj, (int, float)) else p_adj,
            "p_adj_display": format_p_value(p_adj),
            "lower": round(float(lower), 4),
            "upper": round(float(upper), 4),
            "reject": bool(reject),
            "q_stat": round(float(q_stat), 4)
        })

    # Prepare raw data points for boxplot plotting in frontend
    raw_data_points = []
    for name in group_names:
        for val in sanitized_groups[name]:
            raw_data_points.append({
                "Group": name,
                "Value": val
            })

    # Debug details
    debug_details = {
        "group_counts": {name: len(sanitized_groups[name]) for name in group_names},
        "group_means": {name: round(np.mean(sanitized_groups[name]), 4) for name in group_names},
        "group_variances": {name: round(np.var(sanitized_groups[name], ddof=1), 4) for name in group_names},
        "formula_used": formula_used,
        "groups_excluded": excluded_groups,
        "missing_rows_dropped": int(missing_dropped),
        "assumption_test_outputs": {
            "shapiro": shapiro_results,
            "levene": levene_result
        }
    }

    # Calculate compact letter display (CLD)
    group_means_for_cld = {name: np.mean(sanitized_groups[name]) for name in group_names}
    tukey_letters = get_compact_letter_display(group_names, tukey_table, group_means_for_cld)

    # Calculate Control vs Treatment Response Summary
    control_group_name = None
    for name in group_names:
        if format_group_name(name) == "Control":
            control_group_name = name
            break
            
    control_response = []
    if control_group_name is not None:
        control_mean = debug_details["group_means"][control_group_name]
        for name in group_names:
            if name == control_group_name:
                continue
            t_mean = debug_details["group_means"][name]
            diff = t_mean - control_mean
            if control_mean != 0:
                pct_change = (diff / control_mean) * 100
            else:
                pct_change = 0.0
                
            if pct_change <= -10.0:
                interpretation = "Strong growth inhibition"
            elif pct_change < 0.0:
                interpretation = "Slight growth inhibition"
            elif pct_change < 10.0:
                interpretation = "Slight growth improvement"
            else:
                interpretation = "Substantial growth improvement"
                
            control_response.append({
                "treatment": name,
                "diff": round(diff, 4),
                "pct_change": round(pct_change, 2),
                "interpretation": interpretation
            })

    response = {
        "status": "success",
        "factor": factor,
        "crop": crop,
        "variable": variable,
        "day": day,
        "formula": formula_used,
        "summary_stats": summary_stats,
        "anova_table": anova_table,
        "eta_squared": round(eta_squared, 4),
        "eta_interpretation": eta_interpretation,
        "shapiro_results": shapiro_results,
        "levene_result": levene_result,
        "tukey_results": tukey_table,
        "tukey_letters": tukey_letters,
        "control_response": control_response if control_group_name is not None else None,
        "raw_data_points": raw_data_points,
        "alpha": alpha,
        "debug_details": debug_details
    }
    
    return jsonify(response)

def calculate_simple_main_effects(combined_df, group_factor, compare_factor, mse_full, df_err, alpha):
    import scipy.stats as stats
    import numpy as np
    
    results = {}
    letters_dict = {}
    
    # Unique groups of the grouping factor
    unique_groups = sorted(combined_df[group_factor].unique().tolist())
        
    for g in unique_groups:
        # Determine group key string for the dictionary
        if group_factor == "Concentration":
            g_key = "Control" if g == 0.0 else f"{g} g/L"
        else:
            g_key = str(g)
            
        results[g_key] = []
        
        # Filter data for this group
        df_g = combined_df[combined_df[group_factor] == g]
        
        # Unique levels of the comparison factor in this group
        compare_levels = sorted(df_g[compare_factor].unique().tolist())
        k_levels = len(compare_levels)
        
        if k_levels < 2:
            continue
            
        # Calculate critical Q-value using studentized range
        q_crit = stats.studentized_range.ppf(1.0 - alpha, k_levels, df_err)
        
        # Pairwise comparisons
        for i in range(k_levels):
            for j in range(i + 1, k_levels):
                val1 = compare_levels[i]
                val2 = compare_levels[j]
                
                # Retrieve data points for both compared levels
                v1 = df_g[df_g[compare_factor] == val1]["Value"].tolist()
                v2 = df_g[df_g[compare_factor] == val2]["Value"].tolist()
                
                n1 = len(v1)
                n2 = len(v2)
                
                if n1 > 0 and n2 > 0:
                    mean1 = np.mean(v1)
                    mean2 = np.mean(v2)
                    meandiff = mean1 - mean2
                    
                    # Standard error of difference using FULL MODEL MSE
                    se = np.sqrt((mse_full / 2.0) * (1.0 / n1 + 1.0 / n2))
                    if se == 0:
                        se = 1e-6  # safeguard
                        
                    q_stat = abs(meandiff) / se
                    p_adj = stats.studentized_range.sf(q_stat, k_levels, df_err)
                    
                    # CI bounds
                    ci_half = q_crit * se
                    lower = meandiff - ci_half
                    upper = meandiff + ci_half
                    reject = bool(p_adj < alpha)
                    
                    # Format group names for display
                    if compare_factor == "Concentration":
                        lbl1 = "Control" if val1 == 0.0 else f"{val1} g/L"
                        lbl2 = "Control" if val2 == 0.0 else f"{val2} g/L"
                    else:
                        lbl1 = str(val1)
                        lbl2 = str(val2)
                        
                    results[g_key].append({
                        "comparison": f"{lbl1} vs {lbl2}",
                        "group1": lbl1,
                        "group2": lbl2,
                        "meandiff": round(float(meandiff), 4),
                        "p_adj": round(float(p_adj), 6),
                        "lower": round(float(lower), 4),
                        "upper": round(float(upper), 4),
                        "reject": reject
                    })
                    
        # Calculate CLD letters for this slice
        group_means = {}
        groups_list = []
        for val in compare_levels:
            if compare_factor == "Concentration":
                lbl = "Control" if val == 0.0 else f"{val} g/L"
            else:
                lbl = str(val)
            v_vals = df_g[df_g[compare_factor] == val]["Value"].tolist()
            if len(v_vals) > 0:
                group_means[lbl] = float(np.mean(v_vals))
                groups_list.append(lbl)
                
        cld_letters = get_compact_letter_display(groups_list, results[g_key], group_means)
        letters_dict[g_key] = cld_letters
                    
    # Prepare selector values (keys of results that actually have comparisons)
    selector_values = sorted([k for k, v in results.items() if len(v) > 0])
    # If group factor is Concentration, we sort them numerically with Control first
    if group_factor == "Concentration":
        def sort_key(label):
            if label == "Control":
                return 0.0
            try:
                return float(label.replace(" g/L", "").strip())
            except ValueError:
                return 9999.0
        selector_values = sorted(selector_values, key=sort_key)
        
    return {
        "selector_type": group_factor.lower(),
        "selector_values": selector_values,
        "results": results,
        "letters": letters_dict
    }

def run_two_way_analysis(crop, variable, day, selected_biochars_raw, control_mode, alpha_input):
    import scipy.stats as stats
    import numpy as np
    import statsmodels.api as sm
    from statsmodels.formula.api import ols
    import pandas as pd

    alpha = validate_alpha(alpha_input)

    # Filter master dataframe efficiently in one pass
    df_filtered = FLAT_DF[
        (FLAT_DF["Crop"] == crop) &
        (FLAT_DF["Variable"] == variable) &
        (FLAT_DF["Day"] == day)
    ].copy()

    if df_filtered.empty:
        raise ValueError(f"No data found for Crop={crop}, Variable={variable}, Day={day}")

    # Separate Control and Treatment biochars
    controls = df_filtered[df_filtered["Biochar"] == "Control"].copy()
    treatments = df_filtered[df_filtered["Biochar"] != "Control"].copy()

    if treatments.empty:
        raise ValueError("No treatment biochar species found for Two-Way ANOVA.")

    # Retrieve selected biochars parameter
    selected_biochars = []
    if selected_biochars_raw:
        if isinstance(selected_biochars_raw, list):
            selected_biochars = selected_biochars_raw
        else:
            selected_biochars = [b.strip() for b in selected_biochars_raw.split(",") if b.strip()]
        if selected_biochars:
            treatments = treatments[treatments["Biochar"].isin(selected_biochars)]

    unique_biochars = sorted(treatments["Biochar"].unique().tolist())

    if len(unique_biochars) < 2:
        raise ValueError(f"Insufficient treatment biochar species selected (minimum 2 required). Found: {len(unique_biochars)}")
    
    # Parse control handling mode
    if control_mode not in ["replicated", "exclude"]:
        control_mode = "replicated"

    if control_mode == "exclude":
        combined_df = treatments.dropna(subset=["Value"])
        control_mode_label = "Exclude Shared Control"
    else:
        # Replicate Control (concentration 0.0) into each treatment biochar
        replicated_controls = []
        for b in unique_biochars:
            ctrl_copy = controls.copy()
            ctrl_copy["Biochar"] = b
            replicated_controls.append(ctrl_copy)

        combined_df = pd.concat([treatments] + replicated_controls, ignore_index=True)
        combined_df = combined_df.dropna(subset=["Value"])
        control_mode_label = "Include Independent Controls (Default)"
    
    # Ensure Concentration is treated as a float/numeric and then categorical
    combined_df["Concentration"] = combined_df["Concentration"].astype(float)
    
    # Ensure we have enough levels to fit Two-Way ANOVA
    unique_concs = sorted(combined_df["Concentration"].unique().tolist())
    n_unique_concs = len(unique_concs)
    n_unique_biochars = len(combined_df["Biochar"].unique())
    
    if n_unique_concs < 2 or n_unique_biochars < 2:
        raise ValueError(f"Insufficient factor levels for Two-Way ANOVA. Concentrations count: {n_unique_concs}, Biochar Species count: {n_unique_biochars}")

    # Fit OLS model with explicit Sum contrasts for both Biochar and Concentration
    model_formula = 'Value ~ C(Biochar, Sum) * C(Concentration, Sum)'
    try:
        model = ols(model_formula, data=combined_df).fit()
    except Exception as fit_err:
        raise ValueError(f"Failed to fit the Two-Way ANOVA model: {str(fit_err)}. This usually occurs due to insufficient replication or collinearity in the selected data.")
    
    if model.df_resid <= 0:
        raise ValueError("Insufficient replication to perform Two-Way ANOVA. The experimental design leaves 0 degrees of freedom for residuals. Please select more biochar species or check your replication settings.")
    
    # Compute Type III ANOVA
    anova_table = sm.stats.anova_lm(model, typ=3)
    
    # Compute MS and format tables
    table_dict = {}
    for idx, row in anova_table.iterrows():
        ss = row["sum_sq"]
        df = row["df"]
        ms = ss / df if df > 0 else 0
        f_val = row["F"] if "F" in row and not pd.isna(row["F"]) else None
        p_val = row["PR(>F)"] if "PR(>F)" in row and not pd.isna(row["PR(>F)"]) else None
        
        table_dict[idx] = {
            "SS": round(float(ss), 4),
            "df": int(df),
            "MS": round(float(ms), 4),
            "F": round(float(f_val), 4) if f_val is not None else "-",
            "p_value": round(float(p_val), 6) if p_val is not None else "-"
        }
        
    # Degrees of freedom validation check
    N = len(combined_df)
    df_total_expected = N - 1
    df_A = table_dict.get("C(Biochar, Sum)", {}).get("df", 0)
    df_B = table_dict.get("C(Concentration, Sum)", {}).get("df", 0)
    df_AB = table_dict.get("C(Biochar, Sum):C(Concentration, Sum)", {}).get("df", 0)
    df_error = table_dict.get("Residual", {}).get("df", 0)
    df_sum = df_A + df_B + df_AB + df_error
    df_verified = (df_sum == df_total_expected)
    
    # Interaction significance
    p_val_AB = anova_table.loc["C(Biochar, Sum):C(Concentration, Sum)", "PR(>F)"]
    interaction_significant = bool(p_val_AB < alpha) if not pd.isna(p_val_AB) else False

    # Shapiro-Wilk Test
    residuals = model.resid
    n_residuals = len(residuals)
    if n_residuals >= 3:
        try:
            import warnings
            with warnings.catch_warnings(record=True) as w:
                warnings.simplefilter("always")
                sh_stat, sh_p = stats.shapiro(residuals)
            
            shapiro_result = {
                "statistic": round(float(sh_stat), 4),
                "p_value": round(float(sh_p), 6),
                "normal": bool(sh_p >= alpha),
                "note": None
            }
            if len(w) > 0 and any("p-value may not be accurate" in str(warn.message) for warn in w):
                shapiro_result["note"] = "SciPy warning: p-value may not be accurate for very large samples."
        except Exception as sh_err:
            shapiro_result = {
                "statistic": None,
                "p_value": None,
                "normal": None,
                "note": f"Shapiro-Wilk error: {str(sh_err)}"
            }
    else:
        shapiro_result = {
            "statistic": None,
            "p_value": None,
            "normal": None,
            "note": f"N < 3 (only {n_residuals} residuals available)"
        }

    # Group data in a single pass to collect statistics and Levene's groups
    stats_dict = {}
    levene_groups = []
    for (b, c), cell_gp in combined_df.groupby(["Biochar", "Concentration"]):
        vals = cell_gp["Value"].dropna().tolist()
        n = len(vals)
        if n > 0:
            mean = float(np.mean(vals))
            sd = float(np.std(vals, ddof=1)) if n > 1 else 0.0
            stats_dict[(b, float(c))] = {"N": n, "Mean": mean, "SD": sd}
            levene_groups.append(vals)

    # Levene's Test
    if len(levene_groups) >= 2:
        try:
            lev_stat, lev_p = stats.levene(*levene_groups, center='median')
            levene_result = {
                "statistic": round(float(lev_stat), 4),
                "p_value": round(float(lev_p), 6),
                "equal_variance": bool(lev_p >= alpha),
                "note": None
            }
        except Exception as lev_err:
            levene_result = {
                "statistic": None,
                "p_value": None,
                "equal_variance": None,
                "note": f"Levene's error: {str(lev_err)}"
            }
    else:
        levene_result = {
            "statistic": None,
            "p_value": None,
            "equal_variance": None,
            "note": "Insufficient groups (at least 2 groups required)"
        }

    # Cell Means Grid
    cell_means = []
    for b in unique_biochars:
        bio_row = {"Biochar": b}
        for c in unique_concs:
            cf = float(c)
            cell_info = stats_dict.get((b, cf))
            if cell_info:
                bio_row[str(c)] = {
                    "N": cell_info["N"],
                    "Mean": round(cell_info["Mean"], 4),
                    "SD": round(cell_info["SD"], 4)
                }
            else:
                bio_row[str(c)] = {"N": 0, "Mean": "-", "SD": "-"}
        cell_means.append(bio_row)

    # Simple Main Effects
    mse_full = table_dict.get("Residual", {}).get("MS", 0)
    df_err = table_dict.get("Residual", {}).get("df", 0)

    sme_within_biochar = calculate_simple_main_effects(combined_df, "Biochar", "Concentration", mse_full, df_err, alpha)
    sme_within_concentration = calculate_simple_main_effects(combined_df, "Concentration", "Biochar", mse_full, df_err, alpha)

    simple_main_effects = {
        "within_biochar": sme_within_biochar,
        "within_concentration": sme_within_concentration
    }
    
    posthoc_results = sme_within_biochar["results"]

    # Debug details
    debug_details = {
        "factor_levels": {
            "Biochar": unique_biochars,
            "Concentration": unique_concs
        },
        "cell_sample_sizes": {
            b: {str(c): len(combined_df[(combined_df["Biochar"] == b) & (combined_df["Concentration"] == c)]) for c in unique_concs}
            for b in unique_biochars
        },
        "dropped_rows": int(df_filtered["Value"].isnull().sum()),
        "interaction_significant": interaction_significant,
        "model_formula": model_formula,
        "mse_used_for_posthoc": round(float(mse_full), 6),
        "df_error_used_for_posthoc": int(df_err),
        "df_verification": {
            "df_Biochar": df_A,
            "df_Concentration": df_B,
            "df_Interaction": df_AB,
            "df_Residual": df_error,
            "df_Sum": int(df_sum),
            "N_minus_1": int(df_total_expected),
            "df_verified": bool(df_verified)
        }
    }

    # Interaction plot means
    interaction_plot_data = {}
    for b in unique_biochars:
        interaction_plot_data[b] = []
        for c in unique_concs:
            cf = float(c)
            cell_info = stats_dict.get((b, cf))
            if cell_info:
                interaction_plot_data[b].append({
                    "x": c,
                    "y": round(cell_info["Mean"], 4)
                })

    response = {
        "status": "success",
        "crop": crop,
        "variable": variable,
        "day": day,
        "control_mode": control_mode,
        "control_mode_label": control_mode_label,
        "anova_table": table_dict,
        "df_verified": bool(df_verified),
        "interaction_significant": interaction_significant,
        "shapiro_results": shapiro_result,
        "levene_result": levene_result,
        "cell_means": cell_means,
        "posthoc_results": posthoc_results,
        "simple_main_effects": simple_main_effects,
        "interaction_plot_data": interaction_plot_data,
        "alpha": alpha,
        "debug_details": debug_details
    }
    
    return response, combined_df

@app.route("/api/two-way", methods=["POST", "GET"])
def api_two_way():
    if FLAT_DF is None:
        return jsonify({"status": "error", "message": "Data pipeline not initialized"}), 500

    data = request.args if request.method == "GET" else request.get_json(silent=True)
    if not data:
        data = request.form

    crop = data.get("crop")
    variable = data.get("variable")
    day = data.get("day")
    alpha = data.get("alpha", "0.05")
    selected_biochars_raw = data.get("biochars")
    control_mode = data.get("control_mode", "replicated")

    try:
        response, _ = run_two_way_analysis(crop, variable, day, selected_biochars_raw, control_mode, alpha)
        return jsonify(response)
    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)}), 400
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"status": "error", "message": f"Analysis failed: {str(e)}"}), 500

@app.route("/api/export-excel", methods=["POST"])
def export_excel():
    try:
        def format_group_name(name):
            name_str = str(name).strip()
            if name_str == "0.0" or name_str == "0":
                return "Control"
            return name_str

        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "No data provided"}), 400

        crop = data.get("crop", "N/A")
        variable = data.get("variable", "N/A")
        day = data.get("day", "N/A")
        factor = data.get("factor", "N/A")
        biochar = data.get("biochar", "N/A")
        alpha_val = validate_alpha(data.get("alpha", 0.05))

        summary_stats = data.get("summary_stats", [])
        anova_table = data.get("anova_table", {})
        levene_result = data.get("levene_result", {})
        shapiro_results = data.get("shapiro_results", [])
        tukey_results = data.get("tukey_results", [])
        inference_summary = data.get("inference_summary", "N/A")
        chart_image = data.get("chart_image", "")

        # Decode base64 chart image if present
        img = None
        if chart_image and "," in chart_image:
            try:
                header, encoded = chart_image.split(",", 1)
                decoded = base64.b64decode(encoded)
                img = Image(io.BytesIO(decoded))
            except Exception as img_err:
                print(f"Error decoding chart image: {str(img_err)}")

        # Create Workbook
        wb = openpyxl.Workbook()

        # Styles
        font_title = Font(name="Arial", size=14, bold=True, color="1B5E20") # Academic forest green
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=11, bold=True)
        font_regular = Font(name="Arial", size=10)

        fill_header = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid") # Dark gray header
        fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") # subtle light background
        fill_sig = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid") # soft pink/red for significance

        thin_side = Side(border_style="thin", color="D3D3D3")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.views.sheetView[0].showGridLines = True

        ws1.cell(row=1, column=1, value="One-Way ANOVA Research Report Summary").font = font_title
        ws1.row_dimensions[1].height = 25

        ws1.cell(row=3, column=1, value="Field").font = font_header
        ws1.cell(row=3, column=1).fill = fill_header
        ws1.cell(row=3, column=1).alignment = align_center

        ws1.cell(row=3, column=2, value="Value").font = font_header
        ws1.cell(row=3, column=2).fill = fill_header
        ws1.cell(row=3, column=2).alignment = align_center

        between = anova_table.get("Between", {})
        f_stat = between.get("F", "N/A")
        p_val = between.get("p_value", "N/A")
        significant = anova_table.get("Significant", False)

        if isinstance(f_stat, (int, float)):
            f_stat = round(f_stat, 4)
            
        p_val_display = format_p_value(p_val) if p_val is not None and p_val != "N/A" else "N/A"

        eta_squared = data.get("eta_squared", "N/A")
        eta_interpretation = data.get("eta_interpretation", "N/A")
        control_response = data.get("control_response", None)

        summary_data = [
            ("Crop", crop),
            ("Variable", variable),
            ("Day", day),
            ("Grouping Factor", factor),
            ("Biochar Group", biochar),
            ("F-statistic", f_stat),
            ("p-value", p_val_display),
            (f"Significant (p < {alpha_val})", "Yes" if significant else "No"),
            ("Effect Size (Eta Squared)", eta_squared),
            ("Effect Size Interpretation", eta_interpretation)
        ]

        for row_idx, (field, val) in enumerate(summary_data, 4):
            c_f = ws1.cell(row=row_idx, column=1, value=field)
            c_f.font = font_bold
            c_f.border = border_all
            c_f.fill = fill_light

            c_v = ws1.cell(row=row_idx, column=2, value=val)
            c_v.font = font_regular
            c_v.border = border_all
            c_v.alignment = align_left

        # Analysis Settings block
        ws1.merge_cells(start_row=3, start_column=4, end_row=3, end_column=5)
        c_set_title = ws1.cell(row=3, column=4, value="Analysis Settings")
        c_set_title.font = font_header
        c_set_title.fill = fill_header
        c_set_title.alignment = align_center

        ws1.cell(row=3, column=5).border = border_all
        c_set_title.border = border_all

        settings_data = [
            ("Significance Level (α)", alpha_val),
            ("Test", "One-Way ANOVA"),
            ("Post Hoc", "Tukey HSD")
        ]

        for idx, (field, val) in enumerate(settings_data):
            row_idx = 4 + idx
            c_f = ws1.cell(row=row_idx, column=4, value=field)
            c_f.font = font_bold
            c_f.border = border_all
            c_f.fill = fill_light

            c_v = ws1.cell(row=row_idx, column=5, value=val)
            c_v.font = font_regular
            c_v.border = border_all
            c_v.alignment = align_left

        ws1.cell(row=15, column=1, value="Inference Summary:").font = font_bold
        c_inf = ws1.cell(row=16, column=1, value=inference_summary)
        c_inf.font = font_regular
        c_inf.alignment = Alignment(wrap_text=True, vertical="top")
        ws1.merge_cells(start_row=16, start_column=1, end_row=18, end_column=5)

        # Control vs Treatment Response Summary Table
        if control_response:
            ws1.cell(row=20, column=1, value="Control vs Treatment Response Summary").font = font_bold
            
            headers = ["Treatment", "Δ Mean vs Control", "% Change", "Interpretation"]
            for col_num, h_text in enumerate(headers, 1):
                cell = ws1.cell(row=21, column=col_num, value=h_text)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                
            for idx, r in enumerate(control_response):
                r_idx = 22 + idx
                ws1.cell(row=r_idx, column=1, value=format_group_name(r.get("treatment"))).font = font_bold
                ws1.cell(row=r_idx, column=2, value=r.get("diff")).alignment = align_right
                
                pct = r.get("pct_change")
                pct_str = f"{pct:+.1f}%" if isinstance(pct, (int, float)) else str(pct)
                ws1.cell(row=r_idx, column=3, value=pct_str).alignment = align_right
                
                ws1.cell(row=r_idx, column=4, value=r.get("interpretation")).alignment = align_left
                
                for col_num in range(1, 5):
                    cell = ws1.cell(row=r_idx, column=col_num)
                    cell.border = border_all
                    if col_num != 1:
                        cell.font = font_regular

        # Sheet 2: Descriptive Statistics
        ws2 = wb.create_sheet(title="Descriptive Statistics")
        ws2.views.sheetView[0].showGridLines = True
        ws2.cell(row=1, column=1, value="Descriptive Statistics Summary").font = font_title

        desc_headers = ["Group", "N", "Sum (\u03a3X)", "Mean", "Variance", "SD", "SE", "Sum Sq (\u03a3X\u00b2)"]
        for col_num, header in enumerate(desc_headers, 1):
            cell = ws2.cell(row=3, column=col_num, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        for row_idx, r_data in enumerate(summary_stats, 4):
            ws2.cell(row=row_idx, column=1, value=format_group_name(r_data.get("Group"))).font = font_bold
            ws2.cell(row=row_idx, column=2, value=r_data.get("N")).alignment = align_center
            ws2.cell(row=row_idx, column=3, value=r_data.get("Sum"))
            ws2.cell(row=row_idx, column=4, value=r_data.get("Mean"))
            ws2.cell(row=row_idx, column=5, value=r_data.get("Variance"))
            ws2.cell(row=row_idx, column=6, value=r_data.get("SD"))
            
            # Retrieve or calculate SE
            se_val = r_data.get("SE")
            if se_val is None:
                n_val = r_data.get("N", 0)
                sd_val = r_data.get("SD", 0)
                se_val = sd_val / (n_val ** 0.5) if n_val > 0 else 0
            
            ws2.cell(row=row_idx, column=7, value=round(float(se_val), 4))
            ws2.cell(row=row_idx, column=8, value=r_data.get("SumSq"))

            for col_num in range(1, 9):
                cell = ws2.cell(row=row_idx, column=col_num)
                cell.font = font_bold if col_num == 1 else font_regular
                cell.border = border_all
                if col_num > 1:
                    cell.alignment = align_right if col_num != 2 else align_center

        # Sheet 3: ANOVA Table
        ws3 = wb.create_sheet(title="ANOVA Table")
        ws3.views.sheetView[0].showGridLines = True
        ws3.cell(row=1, column=1, value="One-Way ANOVA Table (F-Test)").font = font_title

        anova_headers = ["Source of Variation", "SS", "df", "MS", "F-value", "p-value", "Sig."]
        for col_num, header in enumerate(anova_headers, 1):
            cell = ws3.cell(row=3, column=col_num, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        between_data = anova_table.get("Between", {})
        within_data = anova_table.get("Within", {})
        total_data = anova_table.get("Total", {})

        sig_star = "ns"
        p_val_num = between_data.get("p_value", 1.0)
        if isinstance(p_val_num, (int, float)):
            if p_val_num < 0.001: sig_star = "***"
            elif p_val_num < 0.01: sig_star = "**"
            elif p_val_num < 0.05: sig_star = "*"

        p_val_formatted = format_p_value(p_val_num) if p_val_num is not None and p_val_num != "N/A" else "N/A"

        anova_rows = [
            ("Between Groups (Treatment)", between_data.get("SS"), between_data.get("df"), between_data.get("MS"), between_data.get("F"), p_val_formatted, sig_star),
            ("Within Groups (Error)", within_data.get("SS"), within_data.get("df"), within_data.get("MS"), "", "", ""),
            ("Total", total_data.get("SS"), total_data.get("df"), "", "", "", "")
        ]

        for row_idx, r_data in enumerate(anova_rows, 4):
            for col_num, val in enumerate(r_data, 1):
                cell = ws3.cell(row=row_idx, column=col_num, value=val)
                cell.border = border_all
                if row_idx == 6:  # Total row
                    cell.fill = fill_light
                    cell.font = font_bold
                else:
                    cell.font = font_bold if col_num == 1 else font_regular

                if col_num in [2, 3, 4, 5, 6]:
                    if isinstance(val, (int, float)):
                        cell.value = round(val, 4)
                    cell.alignment = align_right
                elif col_num == 7:
                    cell.alignment = align_center

        # Sheet 4: Assumption Tests
        ws4 = wb.create_sheet(title="Assumption Tests")
        ws4.views.sheetView[0].showGridLines = True

        ws4.cell(row=1, column=1, value="Homogeneity of Variance (Levene's Test)").font = font_title

        ws4.cell(row=3, column=1, value="Levene Statistic").font = font_bold
        ws4.cell(row=3, column=2, value=levene_result.get("Statistic")).alignment = align_right
        ws4.cell(row=4, column=1, value="p-value").font = font_bold
        
        lev_p_raw = levene_result.get("p_value")
        lev_p_formatted = format_p_value(lev_p_raw) if lev_p_raw is not None and lev_p_raw != "N/A" else "N/A"
        ws4.cell(row=4, column=2, value=lev_p_formatted).alignment = align_right
        ws4.cell(row=5, column=1, value="Assumption Met?").font = font_bold
        ws4.cell(row=5, column=2, value="Yes" if levene_result.get("Equal_Variance") else "No").alignment = align_center

        for r in range(3, 6):
            ws4.cell(row=r, column=1).border = border_all
            ws4.cell(row=r, column=1).fill = fill_light
            ws4.cell(row=r, column=2).border = border_all

        ws4.cell(row=7, column=1, value="Levene's Test Interpretation:").font = font_bold
        equal_var = levene_result.get("Equal_Variance", False)
        lev_interp = "Variances appear sufficiently equal across treatment groups. Standard ANOVA assumptions are met." if equal_var else "Group variances differ significantly. ANOVA is generally robust to moderate variance differences when group sizes are equal, but results should be interpreted cautiously."
        c_lev_text = ws4.cell(row=8, column=1, value=lev_interp)
        c_lev_text.font = font_regular
        c_lev_text.alignment = Alignment(wrap_text=True, vertical="top")
        ws4.merge_cells(start_row=8, start_column=1, end_row=9, end_column=5)

        ws4.cell(row=11, column=1, value="Normality Check (Shapiro-Wilk Test)").font = font_title

        shapiro_headers = ["Group", "W Statistic", "p-value", "Interpretation"]
        for col_num, header in enumerate(shapiro_headers, 1):
            cell = ws4.cell(row=13, column=col_num, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        for row_idx, r_data in enumerate(shapiro_results, 14):
            is_normal = r_data.get("Normal")
            if is_normal is None:
                normal_text = r_data.get("Note", "N < 3")
            else:
                normal_text = "Approximately normal" if is_normal else "Possible deviation from normality"

            ws4.cell(row=row_idx, column=1, value=format_group_name(r_data.get("Group"))).font = font_bold
            
            w_stat = r_data.get("Statistic")
            ws4.cell(row=row_idx, column=2, value=round(w_stat, 4) if isinstance(w_stat, (int, float)) else w_stat).alignment = align_right
            
            p_val_sh = r_data.get("p_value")
            p_val_sh_formatted = format_p_value(p_val_sh) if p_val_sh is not None and p_val_sh != "N/A" else "N/A"
            ws4.cell(row=row_idx, column=3, value=p_val_sh_formatted).alignment = align_right
            
            ws4.cell(row=row_idx, column=4, value=normal_text).alignment = align_left

            for col_num in range(1, 5):
                cell = ws4.cell(row=row_idx, column=col_num)
                cell.font = font_bold if col_num == 1 else font_regular
                cell.border = border_all

        # Sheet 5: Tukey HSD
        ws5 = wb.create_sheet(title="Tukey HSD")
        ws5.views.sheetView[0].showGridLines = True
        ws5.cell(row=1, column=1, value="Post-Hoc Pairwise Comparisons (Tukey HSD)").font = font_title

        tukey_headers = ["Comparison", "Mean Difference", "Q Statistic", "Adjusted p-value", "95% CI Lower", "95% CI Upper", "Significant?"]
        for col_num, header in enumerate(tukey_headers, 1):
            cell = ws5.cell(row=3, column=col_num, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center

        for row_idx, r_data in enumerate(tukey_results, 4):
            comp_name = f"{format_group_name(r_data.get('group1'))} vs {format_group_name(r_data.get('group2'))}"
            ws5.cell(row=row_idx, column=1, value=comp_name).font = font_bold
            ws5.cell(row=row_idx, column=2, value=round(r_data.get("meandiff", 0), 4)).alignment = align_right
            
            q_stat_val = r_data.get("q_stat", 0.0)
            ws5.cell(row=row_idx, column=3, value=round(q_stat_val, 4) if isinstance(q_stat_val, (int, float)) else q_stat_val).alignment = align_right
            
            p_adj_val = r_data.get("p_adj")
            p_adj_formatted = format_p_value(p_adj_val) if p_adj_val is not None and p_adj_val != "N/A" else "N/A"
            ws5.cell(row=row_idx, column=4, value=p_adj_formatted).alignment = align_right
            
            ws5.cell(row=row_idx, column=5, value=round(r_data.get("lower", 0), 4)).alignment = align_right
            ws5.cell(row=row_idx, column=6, value=round(r_data.get("upper", 0), 4)).alignment = align_right

            reject = r_data.get("reject", False)
            sig_text = "Significant" if reject else "Not Significant"
            ws5.cell(row=row_idx, column=7, value=sig_text).alignment = align_center

            for col_num in range(1, 8):
                cell = ws5.cell(row=row_idx, column=col_num)
                cell.font = font_bold if col_num == 1 else font_regular
                cell.border = border_all
                if col_num == 7 and reject:
                    cell.fill = fill_sig

        # Add Significance Letter Grouping Table
        tukey_letters = data.get("tukey_letters", {})
        if tukey_letters:
            group_means_dict = {}
            for r_data in summary_stats:
                g_name = format_group_name(r_data.get("Group"))
                group_means_dict[g_name] = r_data.get("Mean", 0.0)
                
            sorted_groups_for_cld = sorted(list(tukey_letters.keys()), key=lambda g: group_means_dict.get(g, 0.0), reverse=True)
            
            start_row_cld = 3 + len(tukey_results) + 3
            ws5.cell(row=start_row_cld - 1, column=1, value="Significance Letter Groupings (Tukey CLD)").font = font_bold
            
            cld_headers = ["Group", "Mean", "Significance Group"]
            for col_num, header in enumerate(cld_headers, 1):
                cell = ws5.cell(row=start_row_cld, column=col_num, value=header)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                
            for idx, g_name in enumerate(sorted_groups_for_cld):
                r_idx = start_row_cld + 1 + idx
                ws5.cell(row=r_idx, column=1, value=g_name).font = font_bold
                ws5.cell(row=r_idx, column=2, value=group_means_dict.get(g_name, 0.0)).alignment = align_right
                
                c_let = ws5.cell(row=r_idx, column=3, value=tukey_letters.get(g_name, ""))
                c_let.alignment = align_center
                c_let.font = Font(name="Arial", size=10, bold=True, color="212529")
                
                for col_num in range(1, 4):
                    cell = ws5.cell(row=r_idx, column=col_num)
                    cell.border = border_all
                    if col_num != 3:
                        cell.font = font_bold if col_num == 1 else font_regular

        # Sheet 6: Graph
        ws6 = wb.create_sheet(title="Graph")
        ws6.views.sheetView[0].showGridLines = True
        ws6.cell(row=1, column=1, value="Seedling Growth Analysis Graph").font = font_title
        if img:
            ws6.add_image(img, "B3")

        # Auto-adjust column widths safely
        for ws in wb.worksheets:
            if ws.title == "Graph":
                continue
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Skip merged cells and long descriptions from auto-width to prevent huge columns
                        if len(val_str) > 50:
                            continue
                        lines = val_str.split('\n')
                        for line in lines:
                            max_len = max(max_len, len(line))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Dynamic filenames
        safe_crop = str(crop).replace(" ", "")
        safe_var = str(variable).replace(" ", "")
        safe_day = str(day).replace(" ", "")
        
        factor_suffix = ""
        if factor == "Concentration":
            factor_suffix = str(biochar).replace(" ", "_")
        elif factor == "Biochar":
            factor_suffix = "Biochar"
        else:
            factor_suffix = str(factor).replace(" ", "")
            
        filename = f"{safe_crop}_{safe_var}_{safe_day}_{factor_suffix}_Report.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"status": "error", "message": f"Excel generation failed: {str(e)}"}), 500

@app.route("/api/export-excel-twoway", methods=["POST"])
def export_excel_twoway():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"status": "error", "message": "No parameters provided"}), 400

        crop = data.get("crop")
        variable = data.get("variable")
        day = data.get("day")
        alpha = data.get("alpha", "0.05")
        selected_biochars_raw = data.get("biochars")
        control_mode = data.get("control_mode", "replicated")

        # Run internal analysis
        response_dict, combined_df = run_two_way_analysis(crop, variable, day, selected_biochars_raw, control_mode, alpha)
        
        # Build Workbook
        wb = openpyxl.Workbook()
        
        # Styles
        font_title = Font(name="Arial", size=14, bold=True, color="1B5E20")  # Forest Green
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=11, bold=True)
        font_regular = Font(name="Arial", size=10)
        
        fill_header = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
        fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        fill_sig = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        
        thin_side = Side(border_style="thin", color="D3D3D3")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        def clean_p_value_format(p):
            if p is None or p == "-" or p == "N/A":
                return "N/A"
            try:
                p_val = float(p)
            except (ValueError, TypeError):
                return str(p)
            if p_val < 0.0001:
                return "< 0.0001"
            return f"{p_val:.4f}"
            
        def style_cell(cell, font=font_regular, fill=None, border=border_all, alignment=None, num_format=None):
            if font: cell.font = font
            if fill: cell.fill = fill
            if border: cell.border = border
            if alignment: cell.alignment = alignment
            if num_format: cell.number_format = num_format

        # Helper to get unique biochars & concentrations
        unique_biochars = response_dict["debug_details"]["factor_levels"]["Biochar"]
        unique_concs = response_dict["debug_details"]["factor_levels"]["Concentration"]
        
        # 1. Sheet 1: Summary
        ws1 = wb.active
        ws1.title = "Analysis Summary"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1.cell(row=1, column=1, value="Two-Way ANOVA Analysis Summary").font = font_title
        ws1.row_dimensions[1].height = 25
        
        headers = ["Parameter", "Value"]
        for col_num, h in enumerate(headers, 1):
            cell = ws1.cell(row=3, column=col_num, value=h)
            style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)
            
        # Analysis Design description
        design_type = "Treatment-only Factorial" if control_mode == "exclude" else "Complete Factorial"
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        summary_rows = [
            ("Crop", crop),
            ("Variable", variable),
            ("Day", day),
            ("Factor A (Independent)", "Biochar Species"),
            ("Factor B (Independent)", "Concentration"),
            ("Selected Biochar Species", ", ".join(unique_biochars)),
            ("Number of Biochar Species", len(unique_biochars)),
            ("Control Handling Mode", response_dict["control_mode_label"]),
            ("Experimental Design", design_type),
            ("Significance Level (α)", response_dict["alpha"]),
            ("Export Timestamp", timestamp),
            ("Software Version", "v1.7.0")
        ]
        
        for idx, (field, val) in enumerate(summary_rows, 4):
            c_f = ws1.cell(row=idx, column=1, value=field)
            style_cell(c_f, font=font_bold, fill=fill_light)
            
            c_v = ws1.cell(row=idx, column=2, value=val)
            style_cell(c_v, font=font_regular, alignment=align_left)
            if idx % 2 == 1:
                c_v.fill = fill_light

        # Scientific Interpretation block
        p_A = response_dict["anova_table"].get("C(Biochar, Sum)", {}).get("p_value", 1.0)
        p_B = response_dict["anova_table"].get("C(Concentration, Sum)", {}).get("p_value", 1.0)
        p_AB = response_dict["anova_table"].get("C(Biochar, Sum):C(Concentration, Sum)", {}).get("p_value", 1.0)
        
        sigA = p_A != "-" and p_A < float(alpha)
        sigB = p_B != "-" and p_B < float(alpha)
        sigAB = p_AB != "-" and p_AB < float(alpha)
        
        strA = f"Significant (p = {p_A:.4f})" if sigA else (f"Not Significant (p = {p_A:.4f})" if isinstance(p_A, (int, float)) else f"Not Significant ({p_A})")
        strB = f"Significant (p = {p_B:.4f})" if sigB else (f"Not Significant (p = {p_B:.4f})" if isinstance(p_B, (int, float)) else f"Not Significant ({p_B})")
        strAB = f"Significant (p = {p_AB:.4f})" if sigAB else (f"Not Significant (p = {p_AB:.4f})" if isinstance(p_AB, (int, float)) else f"Not Significant ({p_AB})")
        
        summary_text = (
            f"Scientific Inference Summary:\n"
            f"• Factor A (Biochar Species): {strA}. Biochar species differ in their general effect.\n"
            f"• Factor B (Concentration): {strB}. Concentrations differ in their general effect.\n"
            f"• Interaction (Biochar Species × Concentration): {strAB}. The response curves are {'non-parallel' if sigAB else 'parallel'}.\n\n"
        )
        if sigAB:
            summary_text += (
                f"Scientific Interpretation Rule: Because the interaction effect is statistically significant (p < {alpha}), "
                f"you cannot interpret the main effects of Biochar Species or Concentration directly. Focus instead on the "
                f"Post-hoc Analysis of Simple Main Effects (Tukey HSD) comparisons shown in this report."
            )
        else:
            summary_text += (
                f"Scientific Interpretation Rule: Because the interaction effect is not significant, the main effects can be "
                f"interpreted directly. Main effects indicate consistent trends across all groups."
            )
            
        start_row_inf = 4 + len(summary_rows) + 2
        ws1.cell(row=start_row_inf, column=1, value="Scientific Interpretation Summary").font = font_bold
        c_inf = ws1.cell(row=start_row_inf + 1, column=1, value=summary_text)
        style_cell(c_inf, font=font_regular, alignment=Alignment(wrap_text=True, vertical="top"))
        ws1.merge_cells(start_row=start_row_inf + 1, start_column=1, end_row=start_row_inf + 6, end_column=5)
        for r in range(start_row_inf + 1, start_row_inf + 7):
            for c in range(1, 6):
                ws1.cell(row=r, column=c).border = border_all
        ws1.freeze_panes = "A4"

        # 2. Sheet 2: Experimental Design
        ws2 = wb.create_sheet(title="Experimental Design")
        ws2.views.sheetView[0].showGridLines = True
        ws2.cell(row=1, column=1, value="Experimental Design Summary").font = font_title
        
        for col_num, h in enumerate(headers, 1):
            cell = ws2.cell(row=3, column=col_num, value=h)
            style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)
            
        cell_sizes = [len(combined_df[(combined_df["Biochar"] == b) & (combined_df["Concentration"] == c)]) for b in unique_biochars for c in unique_concs]
        if len(cell_sizes) > 0:
            min_size = min(cell_sizes)
            max_size = max(cell_sizes)
            replicates_str = str(min_size) if min_size == max_size else f"{min_size} - {max_size} (Unbalanced Design)"
        else:
            replicates_str = "0"
            
        design_rows = [
            ("Factor A (Independent)", "Biochar Species"),
            ("Factor B (Independent)", "Concentration"),
            ("Active Biochar Species", ", ".join(unique_biochars)),
            ("Active Concentration Levels", ", ".join([("Control" if c == 0.0 else f"{c} g/L") for c in unique_concs])),
            ("Number of Replicates", replicates_str),
            ("Total Observations", len(combined_df)),
            ("Control Handling Mode", response_dict["control_mode_label"])
        ]
        
        for idx, (field, val) in enumerate(design_rows, 4):
            c_f = ws2.cell(row=idx, column=1, value=field)
            style_cell(c_f, font=font_bold, fill=fill_light)
            
            c_v = ws2.cell(row=idx, column=2, value=val)
            style_cell(c_v, font=font_regular, alignment=align_left)
            if idx % 2 == 1:
                c_v.fill = fill_light
        ws2.freeze_panes = "A4"

        # 3. Sheet 3: Cell Means Matrix
        ws3 = wb.create_sheet(title="Cell Means Matrix")
        ws3.views.sheetView[0].showGridLines = True
        ws3.cell(row=1, column=1, value="Cell Replication & Means Matrix (Biochar Species × Concentration)").font = font_title
        
        ws3.cell(row=3, column=1, value="Biochar Species").font = font_header
        ws3.cell(row=3, column=1).fill = fill_header
        ws3.cell(row=3, column=1).alignment = align_center
        ws3.cell(row=3, column=1).border = border_all
        
        for idx, c in enumerate(unique_concs, 2):
            c_lbl = "Control" if c == 0.0 else f"{c} g/L"
            cell = ws3.cell(row=3, column=idx, value=c_lbl)
            style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)
            
        # Data rows
        for row_idx, r_data in enumerate(response_dict["cell_means"], 4):
            ws3.row_dimensions[row_idx].height = 42
            c_bio = ws3.cell(row=row_idx, column=1, value=r_data["Biochar"])
            style_cell(c_bio, font=font_bold, alignment=align_left)
            if row_idx % 2 == 1:
                c_bio.fill = fill_light
                
            for col_idx, c in enumerate(unique_concs, 2):
                c_str = str(float(c))
                cell_info = r_data.get(c_str)
                if not cell_info:
                    c_str_alt = f"{float(c):.1f}"
                    cell_info = r_data.get(c_str_alt)
                if not cell_info:
                    cell_info = r_data.get(str(c))
                    
                cell = ws3.cell(row=row_idx, column=col_idx)
                
                if cell_info and cell_info.get("N", 0) > 0:
                    val_str = f"{cell_info['Mean']:.4f}\nSD: {cell_info['SD']:.4f}\n(N={cell_info['N']})"
                    cell.value = val_str
                    style_cell(cell, font=font_regular, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
                else:
                    cell.value = "-"
                    style_cell(cell, font=font_regular, alignment=align_center)
                    
                if row_idx % 2 == 1:
                    cell.fill = fill_light
        ws3.freeze_panes = "A4"

        # 4. Sheet 4: Type III ANOVA
        ws4 = wb.create_sheet(title="Type III ANOVA")
        ws4.views.sheetView[0].showGridLines = True
        ws4.cell(row=1, column=1, value="Type III ANOVA Table (For Unbalanced Designs)").font = font_title
        
        anova_headers = ["Source of Variation", "Sum Sq (SS)", "df", "Mean Sq (MS)", "F-value", "p-value", "Sig."]
        for col_num, h in enumerate(anova_headers, 1):
            cell = ws4.cell(row=3, column=col_num, value=h)
            style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)
            
        key_labels = {
            "Intercept": "Intercept",
            "C(Biochar, Sum)": "Biochar Species (Factor A)",
            "C(Concentration, Sum)": "Concentration (Factor B)",
            "C(Biochar, Sum):C(Concentration, Sum)": "Biochar Species & Concentration (Interaction)",
            "Residual": "Error (Residuals)"
        }
        
        calculatedTotalSS = 0
        calculatedTotalDf = 0
        
        anova_data = response_dict["anova_table"]
        row_idx = 4
        for key, label in key_labels.items():
            row_vals = anova_data.get(key)
            if not row_vals:
                continue
                
            ss = row_vals["SS"]
            df = row_vals["df"]
            ms = row_vals["MS"]
            f_val = row_vals["F"]
            p_val = row_vals["p_value"]
            
            if key != "Intercept":
                calculatedTotalSS += ss
                calculatedTotalDf += df
                
            sig_star = "ns"
            if isinstance(p_val, (int, float)):
                if p_val < 0.001: sig_star = "***"
                elif p_val < 0.01: sig_star = "**"
                elif p_val < 0.05: sig_star = "*"
                else: sig_star = "ns"
            elif p_val == "-":
                sig_star = "-"
                
            p_formatted = clean_p_value_format(p_val)
            
            ws4.cell(row=row_idx, column=1, value=label).font = font_bold
            ws4.cell(row=row_idx, column=2, value=ss)
            ws4.cell(row=row_idx, column=3, value=df)
            ws4.cell(row=row_idx, column=4, value=ms)
            
            f_cell = ws4.cell(row=row_idx, column=5, value=f_val)
            p_cell = ws4.cell(row=row_idx, column=6, value=p_formatted)
            ws4.cell(row=row_idx, column=7, value=sig_star).alignment = align_center
            
            for c_idx in range(1, 8):
                cell = ws4.cell(row=row_idx, column=c_idx)
                cell.border = border_all
                if c_idx != 1:
                    cell.font = font_regular
                    if c_idx in [2, 3, 4, 5, 6]:
                        cell.alignment = align_right
                        if c_idx in [2, 4] or (c_idx == 5 and isinstance(f_val, (int, float))):
                            cell.number_format = "0.0000"
            row_idx += 1
            
        # Total Row
        ws4.cell(row=row_idx, column=1, value="Total (Corrected)").font = font_bold
        ws4.cell(row=row_idx, column=2, value=calculatedTotalSS)
        ws4.cell(row=row_idx, column=3, value=calculatedTotalDf)
        
        for c_idx in range(1, 8):
            cell = ws4.cell(row=row_idx, column=c_idx)
            cell.border = border_all
            cell.fill = fill_light
            if c_idx != 1:
                cell.font = font_regular
                if c_idx in [2, 3]:
                    cell.alignment = align_right
                    if c_idx == 2:
                        cell.number_format = "0.0000"
                        
        # Append inference summary text below
        start_row_inf_anova = row_idx + 3
        ws4.cell(row=start_row_inf_anova - 1, column=1, value="Scientific Inference Summary").font = font_bold
        c_inf_anova = ws4.cell(row=start_row_inf_anova, column=1, value=summary_text)
        style_cell(c_inf_anova, font=font_regular, alignment=Alignment(wrap_text=True, vertical="top"))
        ws4.merge_cells(start_row=start_row_inf_anova, start_column=1, end_row=start_row_inf_anova + 5, end_column=7)
        for r in range(start_row_inf_anova, start_row_inf_anova + 6):
            for c in range(1, 8):
                ws4.cell(row=r, column=c).border = border_all
        ws4.freeze_panes = "A4"

        # 5. Sheet 5: Assumption Diagnostics
        ws5 = wb.create_sheet(title="Assumption Diagnostics")
        ws5.views.sheetView[0].showGridLines = True
        ws5.cell(row=1, column=1, value="Homogeneity of Variance (Levene's Test)").font = font_title
        
        levene_data = response_dict["levene_result"]
        lev_stat = levene_data.get("statistic")
        lev_p = levene_data.get("p_value")
        lev_met = levene_data.get("equal_variance")
        
        lev_p_formatted = clean_p_value_format(lev_p)
        
        if lev_met is None:
            lev_dec = "N/A"
            lev_interp = levene_data.get("note", "Could not compute Levene's test.")
        else:
            lev_dec = "Assumption Met (Equal Variances)" if lev_met else "Assumption Violated (Variances unequal)"
            lev_interp = "Variances appear sufficiently equal across all Biochar Species × Concentration cells." if lev_met else "Cell variances differ significantly. Two-Way ANOVA is robust to moderate variance differences when sample sizes are balanced, but interpretation should be cautious."
            
        ws5.cell(row=3, column=1, value="Levene Statistic").font = font_bold
        ws5.cell(row=3, column=2, value=lev_stat).alignment = align_right
        ws5.cell(row=3, column=2).number_format = "0.0000"
        
        ws5.cell(row=4, column=1, value="p-value").font = font_bold
        ws5.cell(row=4, column=2, value=lev_p_formatted).alignment = align_right
        
        ws5.cell(row=5, column=1, value="Decision").font = font_bold
        ws5.cell(row=5, column=2, value=lev_dec).alignment = align_left
        
        for r in range(3, 6):
            ws5.cell(row=r, column=1).border = border_all
            ws5.cell(row=r, column=1).fill = fill_light
            ws5.cell(row=r, column=2).border = border_all
            ws5.cell(row=r, column=2).font = font_regular
            
        ws5.cell(row=7, column=1, value="Scientific Interpretation:").font = font_bold
        c_lev_text = ws5.cell(row=8, column=1, value=lev_interp)
        style_cell(c_lev_text, font=font_regular, alignment=Alignment(wrap_text=True, vertical="top"))
        ws5.merge_cells(start_row=8, start_column=1, end_row=9, end_column=5)
        for r in range(8, 10):
            for c in range(1, 6):
                ws5.cell(row=r, column=c).border = border_all
                
        # Shapiro-Wilk Section
        ws5.cell(row=11, column=1, value="Normality of Residuals (Shapiro-Wilk Test)").font = font_title
        
        shapiro_data = response_dict["shapiro_results"]
        sh_stat = shapiro_data.get("statistic")
        sh_p = shapiro_data.get("p_value")
        sh_normal = shapiro_data.get("normal")
        
        sh_p_formatted = clean_p_value_format(sh_p)
        
        if sh_normal is None:
            sh_dec = "N/A"
            sh_interp = shapiro_data.get("note", "Could not compute normality check.")
        else:
            sh_dec = "Assumption Met (Normal Residuals)" if sh_normal else "Assumption Violated (Non-normal residuals)"
            sh_interp = "Standard normality assumptions for the OLS model are met." if sh_normal else "Residuals deviate significantly from a normal distribution. While ANOVA is robust to mild deviations from normality with larger sample sizes, results should be interpreted with caution."
            if shapiro_data.get("note"):
                sh_interp += f"\nNote: {shapiro_data.get('note')}"
                
        ws5.cell(row=13, column=1, value="Shapiro-Wilk Statistic").font = font_bold
        ws5.cell(row=13, column=2, value=sh_stat).alignment = align_right
        ws5.cell(row=13, column=2).number_format = "0.0000"
        
        ws5.cell(row=14, column=1, value="p-value").font = font_bold
        ws5.cell(row=14, column=2, value=sh_p_formatted).alignment = align_right
        
        ws5.cell(row=15, column=1, value="Decision").font = font_bold
        ws5.cell(row=15, column=2, value=sh_dec).alignment = align_left
        
        for r in range(13, 16):
            ws5.cell(row=r, column=1).border = border_all
            ws5.cell(row=r, column=1).fill = fill_light
            ws5.cell(row=r, column=2).border = border_all
            ws5.cell(row=r, column=2).font = font_regular
            
        ws5.cell(row=17, column=1, value="Scientific Interpretation:").font = font_bold
        c_sh_text = ws5.cell(row=18, column=1, value=sh_interp)
        style_cell(c_sh_text, font=font_regular, alignment=Alignment(wrap_text=True, vertical="top"))
        ws5.merge_cells(start_row=18, start_column=1, end_row=19, end_column=5)
        for r in range(18, 20):
            for c in range(1, 6):
                ws5.cell(row=r, column=c).border = border_all
        ws5.freeze_panes = "A4"

        # 6. Sheet 6: Simple Main Effects
        ws6 = wb.create_sheet(title="Simple Main Effects")
        ws6.views.sheetView[0].showGridLines = True
        ws6.cell(row=1, column=1, value="Post-hoc Analysis of Simple Main Effects (Tukey HSD)").font = font_title
        
        sme_data = response_dict["simple_main_effects"]
        
        # Section A: Concentrations within Biochar Species
        ws6.cell(row=3, column=1, value="Section A: Compare Concentrations within Biochar Species").font = font_bold
        
        sme_headers = ["Group", "Comparison", "Mean Difference", "Adjusted p-value", "95% CI Lower", "95% CI Upper", "Significant?"]
        for col_num, h in enumerate(sme_headers, 1):
            cell = ws6.cell(row=4, column=col_num, value=h)
            style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)
            
        row_idx = 5
        wb_results = sme_data.get("within_biochar", {}).get("results", {})
        for g_key, comps in wb_results.items():
            for comp in comps:
                ws6.cell(row=row_idx, column=1, value=g_key).font = font_bold
                comp_text = comp.get("comparison") or f"{comp.get('group1')} vs {comp.get('group2')}"
                ws6.cell(row=row_idx, column=2, value=comp_text)
                
                md = comp.get("meandiff", 0.0)
                ws6.cell(row=row_idx, column=3, value=md)
                
                p_adj = comp.get("p_adj")
                p_adj_formatted = clean_p_value_format(p_adj)
                ws6.cell(row=row_idx, column=4, value=p_adj_formatted)
                
                ws6.cell(row=row_idx, column=5, value=comp.get("lower", 0.0))
                ws6.cell(row=row_idx, column=6, value=comp.get("upper", 0.0))
                
                reject = comp.get("reject", False)
                sig_text = "Significant" if reject else "Not Significant"
                ws6.cell(row=row_idx, column=7, value=sig_text)
                
                # Style row
                for col_idx in range(1, 8):
                    cell = ws6.cell(row=row_idx, column=col_idx)
                    cell.border = border_all
                    if col_idx != 1:
                        cell.font = font_regular
                        if col_idx in [3, 4, 5, 6]:
                            cell.alignment = align_right
                            if col_idx in [3, 5, 6]:
                                cell.number_format = "0.0000"
                        elif col_idx == 7:
                            cell.alignment = align_center
                            if reject:
                                cell.fill = fill_sig
                row_idx += 1
                
        # Section B: Biochar Species within Concentration
        row_idx += 3
        ws6.cell(row=row_idx, column=1, value="Section B: Compare Biochar Species within Concentration").font = font_bold
        row_idx += 1
        
        for col_num, h in enumerate(sme_headers, 1):
            cell = ws6.cell(row=row_idx, column=col_num, value=h)
            style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)
            
        row_idx += 1
        wc_results = sme_data.get("within_concentration", {}).get("results", {})
        for g_key, comps in wc_results.items():
            for comp in comps:
                ws6.cell(row=row_idx, column=1, value=g_key).font = font_bold
                comp_text = comp.get("comparison") or f"{comp.get('group1')} vs {comp.get('group2')}"
                ws6.cell(row=row_idx, column=2, value=comp_text)
                
                md = comp.get("meandiff", 0.0)
                ws6.cell(row=row_idx, column=3, value=md)
                
                p_adj = comp.get("p_adj")
                p_adj_formatted = clean_p_value_format(p_adj)
                ws6.cell(row=row_idx, column=4, value=p_adj_formatted)
                
                ws6.cell(row=row_idx, column=5, value=comp.get("lower", 0.0))
                ws6.cell(row=row_idx, column=6, value=comp.get("upper", 0.0))
                
                reject = comp.get("reject", False)
                sig_text = "Significant" if reject else "Not Significant"
                ws6.cell(row=row_idx, column=7, value=sig_text)
                
                # Style row
                for col_idx in range(1, 8):
                    cell = ws6.cell(row=row_idx, column=col_idx)
                    cell.border = border_all
                    if col_idx != 1:
                        cell.font = font_regular
                        if col_idx in [3, 4, 5, 6]:
                            cell.alignment = align_right
                            if col_idx in [3, 5, 6]:
                                cell.number_format = "0.0000"
                        elif col_idx == 7:
                            cell.alignment = align_center
                            if reject:
                                cell.fill = fill_sig
                row_idx += 1
        ws6.freeze_panes = "A5"

        # 7. Sheet 7: Interaction Plot Data
        ws7 = wb.create_sheet(title="Interaction Plot Data")
        ws7.views.sheetView[0].showGridLines = True
        ws7.cell(row=1, column=1, value="Interaction Plot Data Summary").font = font_title
        
        plot_headers = ["Biochar Species", "Concentration", "Mean", "SD", "N"]
        for col_num, h in enumerate(plot_headers, 1):
            cell = ws7.cell(row=3, column=col_num, value=h)
            style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)
            
        grouped = combined_df.groupby(["Biochar", "Concentration"], as_index=False)["Value"].agg(["mean", "std", "count"])
        grouped = grouped.sort_values(by=["Biochar", "Concentration"])
        
        for idx, r in enumerate(grouped.itertuples(index=False), 4):
            b_val = r.Biochar
            c_val = float(r.Concentration)
            m_val = float(r.mean)
            s_val = float(r.std) if pd.notnull(r.std) else 0.0
            n_val = int(r.count)
            
            ws7.cell(row=idx, column=1, value=b_val).font = font_bold
            ws7.cell(row=idx, column=2, value=c_val)
            ws7.cell(row=idx, column=3, value=m_val)
            ws7.cell(row=idx, column=4, value=s_val)
            ws7.cell(row=idx, column=5, value=n_val)
            
            for c_idx in range(1, 6):
                cell = ws7.cell(row=idx, column=c_idx)
                cell.border = border_all
                if cell.row % 2 == 1:
                    cell.fill = fill_light
                if c_idx != 1:
                    cell.font = font_regular
                    if c_idx in [2, 3, 4]:
                        cell.alignment = align_right
                        if c_idx in [3, 4]:
                            cell.number_format = "0.0000"
                        elif c_idx == 2:
                            cell.number_format = "0.0"
                    elif c_idx == 5:
                        cell.alignment = align_center
        ws7.freeze_panes = "A4"

        # 8. Sheet 8: Analysis Dataset
        ws8 = wb.create_sheet(title="Analysis Dataset")
        ws8.views.sheetView[0].showGridLines = True
        ws8.cell(row=1, column=1, value="Analysis Dataset (OLS Model Inputs)").font = font_title
        
        dataset_headers = ["Biochar Species", "Concentration", "Value"]
        for col_num, h in enumerate(dataset_headers, 1):
            cell = ws8.cell(row=3, column=col_num, value=h)
            style_cell(cell, font=font_header, fill=fill_header, alignment=align_center)
            
        for idx, r in enumerate(combined_df.itertuples(index=False), 4):
            ws8.cell(row=idx, column=1, value=r.Biochar).font = font_bold
            ws8.cell(row=idx, column=2, value=float(r.Concentration))
            ws8.cell(row=idx, column=3, value=float(r.Value))
            
            for c_idx in range(1, 4):
                cell = ws8.cell(row=idx, column=c_idx)
                cell.border = border_all
                if cell.row % 2 == 1:
                    cell.fill = fill_light
                if c_idx != 1:
                    cell.font = font_regular
                    cell.alignment = align_right
                    if c_idx == 2:
                        cell.number_format = "0.0"
                    elif c_idx == 3:
                        cell.number_format = "0.0000"
        ws8.freeze_panes = "A4"

        # Auto-adjust column widths safely (skipping sheet titles)
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.row == 1:
                        continue
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Skip long interpretation sentences from auto-width
                        if len(val_str) > 50:
                            continue
                        lines = val_str.split('\n')
                        for line in lines:
                            max_len = max(max_len, len(line))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Filename formatting
        safe_crop = str(crop).replace(" ", "")
        safe_var = str(variable).replace(" ", "")
        safe_day = str(day).replace(" ", "")
        filename = f"TwoWayANOVA_{safe_crop}_{safe_var}_{safe_day}_{datetime.date.today().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return jsonify({"status": "error", "message": f"Excel generation failed: {str(e)}"}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)

